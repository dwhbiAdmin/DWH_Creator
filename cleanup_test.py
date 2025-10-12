import pandas as pd
from openpyxl import load_workbook

wb_path = '_DWH_Projects/Project_AW_sales/2_workbench/workbench_AW_Sales.xlsx'

# Read current columns
df = pd.read_excel(wb_path, sheet_name='columns')
print(f'Before cleanup: {len(df)} rows')

# Keep only s0 (original data)
df_original = df[df['stage_id'] == 's0'].copy()
print(f'Keeping s0: {len(df_original)} rows')

# Load workbook and clear columns sheet
wb = load_workbook(wb_path)
ws = wb['columns']

# Delete all data rows (keep header in row 1)
for row in range(ws.max_row, 1, -1):
    ws.delete_rows(row)

# Write back header and original data
header = list(df_original.columns)
ws.append(header)

for idx, row_data in df_original.iterrows():
    ws.append([row_data.get(col, '') for col in header])

wb.save(wb_path)
print('Cleanup complete - ready for testing')
