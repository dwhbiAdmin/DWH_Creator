"""
Cascade Operations Module
=========================

This module implements field cascading functionality for the workbench.
It handles the automatic propagation of fields from upstream artifacts to downstream artifacts
based on their relationships and configurations.
"""

import pandas as pd
from pathlib import Path
import sys

# Add src directory to path for imports
current_dir = Path(__file__).parent
src_dir = current_dir.parent.parent.parent
sys.path.insert(0, str(src_dir))

from utils.c_workbench_excel_utils import ExcelUtils
from utils.z_logger import Logger


class CascadeOperations:
    """Handles cascade operations for workbench fields."""
    
    def __init__(self, workbook_path: str):
        """
        Initialize the cascade operations.
        
        Args:
            workbook_path: Path to the workbook Excel file
        """
        self.workbook_path = workbook_path
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
    
    def cascade_main(self):
        """
        Main cascade function that iterates through artifacts and cascades fields.
        
        This function:
        1. Reads the artifacts sheet from row 2
        2. Retrieves artifact information for each row
        3. Looks up stage information from the stages sheet
        4. Passes all information to cascade_engine for processing
        """
        try:
            self.logger.info("Starting cascade_main operation")
            
            # Read artifacts sheet (from row 2, since row 1 is header)
            df_artifacts = pd.read_excel(self.workbook_path, sheet_name='artifacts')
            
            # Read stages sheet for lookup
            df_stages = pd.read_excel(self.workbook_path, sheet_name='stages')
            
            self.logger.info(f"Found {len(df_artifacts)} artifacts to process")
            
            # Iterate through each artifact (from row 2, which is index 0 in DataFrame)
            for idx, row in df_artifacts.iterrows():
                # Retrieve fields from artifact row
                stage_id = row.get('stage_id', '')
                stage_name = row.get('stage_name', '')
                artifact_id = row.get('artifact_id', '')
                artifact_name = row.get('artifact_name', '')
                artifact_type = row.get('artifact_type', '')
                artifact_topology = row.get('artifact_topology', '')
                upstream_artifact = row.get('upstream_artifact', '')
                upstream_relation = row.get('upstream_relation', '')
                
                self.logger.info(f"Processing artifact: {artifact_id} - {artifact_name}")
                
                # Look up stage information from stages sheet
                stage_row = df_stages[df_stages['stage_id'] == stage_id]
                
                if stage_row.empty:
                    self.logger.warning(f"Stage {stage_id} not found in stages sheet, skipping artifact {artifact_id}")
                    continue
                
                # Retrieve additional values from stages sheet
                artifact_side = stage_row.iloc[0].get('artifact_side', '')
                platform = stage_row.iloc[0].get('platform', '')
                
                self.logger.info(f"  Stage: {stage_name}, Side: {artifact_side}, Platform: {platform}")
                
                # Pass all fields to cascade_engine
                self.cascade_engine(
                    stage_id=stage_id,
                    stage_name=stage_name,
                    artifact_id=artifact_id,
                    artifact_name=artifact_name,
                    artifact_type=artifact_type,
                    artifact_topology=artifact_topology,
                    upstream_artifact=upstream_artifact,
                    upstream_relation=upstream_relation,
                    artifact_side=artifact_side,
                    platform=platform
                )
            
            # Clean up duplicates after all cascading is complete
            self.logger.info("Cleaning up duplicate column names in artifacts...")
            self._cleanup_duplicate_columns()
            
            self.logger.info("Cascade_main operation completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error in cascade_main: {str(e)}")
            raise
    
    def cascade_engine(self, stage_id: str, stage_name: str, artifact_id: str, 
                       artifact_name: str, artifact_type: str, artifact_topology: str,
                       upstream_artifact: str, upstream_relation: str, 
                       artifact_side: str, platform: str):
        """
        Engine that cascades fields from upstream artifacts to the current artifact.
        
        Args:
            stage_id: Stage identifier (e.g., 's1')
            stage_name: Stage name (e.g., '1_bronze')
            artifact_id: Current artifact identifier
            artifact_name: Current artifact name
            artifact_type: Type of artifact (dimension, fact, etc.)
            artifact_topology: Topology of artifact
            upstream_artifact: Upstream artifact ID(s) - can be multiple separated by ';'
            upstream_relation: Relation type to upstream
            artifact_side: 'source' or 'business'
            platform: Target platform (e.g., 'databricks', 'power_bi')
        """
        try:
            # Check if there is an upstream artifact
            if not upstream_artifact or pd.isna(upstream_artifact):
                self.logger.info(f"  No upstream artifact for {artifact_id}, skipping")
                return
            
            # Convert to string and handle potential NaN
            upstream_artifact = str(upstream_artifact).strip()
            
            if not upstream_artifact or upstream_artifact == 'nan':
                self.logger.info(f"  No upstream artifact for {artifact_id}, skipping")
                return
            
            # Read necessary sheets
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            df_data_mappings = pd.read_excel(self.workbook_path, sheet_name='conf_4_data_mappings')
            
            # Get the last column_id to determine next ID
            if df_columns.empty or 'column_id' not in df_columns.columns:
                column_last_index = 0
            else:
                # Extract numeric part from column_id (e.g., 'c_34' -> 34)
                last_column_id = df_columns['column_id'].iloc[-1]
                try:
                    column_last_index = int(last_column_id.split('_')[1])
                except (IndexError, ValueError):
                    column_last_index = len(df_columns)
            
            self.logger.info(f"  Last column index: {column_last_index}")
            
            # Split multiple upstream artifacts if separated by ';'
            upstream_artifacts = [ua.strip() for ua in upstream_artifact.split(';') if ua.strip()]
            
            self.logger.info(f"  Processing {len(upstream_artifacts)} upstream artifact(s)")
            
            # Process each upstream artifact
            for upstream_art_id in upstream_artifacts:
                self._process_upstream_artifact(
                    upstream_art_id=upstream_art_id,
                    stage_id=stage_id,
                    stage_name=stage_name,
                    artifact_id=artifact_id,
                    artifact_name=artifact_name,
                    artifact_side=artifact_side,
                    platform=platform,
                    df_columns=df_columns,
                    df_data_mappings=df_data_mappings,
                    column_last_index=column_last_index
                )
            
        except Exception as e:
            self.logger.error(f"Error in cascade_engine for artifact {artifact_id}: {str(e)}")
            raise
    
    def _process_upstream_artifact(self, upstream_art_id: str, stage_id: str, 
                                   stage_name: str, artifact_id: str, artifact_name: str,
                                   artifact_side: str, platform: str, 
                                   df_columns: pd.DataFrame, df_data_mappings: pd.DataFrame,
                                   column_last_index: int):
        """
        Process a single upstream artifact and cascade its fields.
        
        Args:
            upstream_art_id: ID of the upstream artifact
            stage_id: Target stage ID
            stage_name: Target stage name
            artifact_id: Target artifact ID
            artifact_name: Target artifact name
            artifact_side: 'source' or 'business'
            platform: Target platform
            df_columns: DataFrame of existing columns
            df_data_mappings: DataFrame of data type mappings
            column_last_index: Last used column index
        """
        # Check if upstream artifact exists in columns sheet
        df_upstream_fields = df_columns[df_columns['artifact_id'] == upstream_art_id].copy()
        
        if df_upstream_fields.empty:
            self.logger.info(f"    Upstream artifact {upstream_art_id} has no fields, skipping")
            return
        
        self.logger.info(f"    Found {len(df_upstream_fields)} fields in upstream artifact {upstream_art_id}")
        
        # FIX 1: Get existing column names in the target artifact to check for duplicates
        existing_columns_in_artifact = df_columns[df_columns['artifact_id'] == artifact_id]['column_name'].tolist()
        
        # Prepare new rows to insert
        new_rows = []
        
        # FIX 2: Track order counter for attributes starting at 100
        attribute_order_counter = 100
        
        # Process each field from upstream artifact
        for idx, upstream_row in df_upstream_fields.iterrows():
            # Get values from upstream
            upstream_column_name = upstream_row.get('column_name', '')
            upstream_data_type = upstream_row.get('data_type', '')
            upstream_order = upstream_row.get('order', 0)
            upstream_column_business_name = upstream_row.get('column_business_name', '')
            upstream_column_group = upstream_row.get('column_group', '')
            upstream_column_comment = upstream_row.get('column_comment', '')
            
            # Determine what the column_name will be in the target artifact
            if artifact_side == 'business':
                if pd.notna(upstream_column_business_name) and upstream_column_business_name:
                    target_column_name = upstream_column_business_name
                else:
                    target_column_name = upstream_column_name
            else:
                target_column_name = upstream_column_name
            
            # FIX 1: Check if column_name already exists in the target artifact
            if target_column_name in existing_columns_in_artifact:
                self.logger.info(f"      Skipping duplicate column '{target_column_name}' in artifact {artifact_id}")
                continue
            
            # Add to existing columns list to prevent duplicates within this batch
            existing_columns_in_artifact.append(target_column_name)
            
            # Increment column index for new column
            column_last_index += 1
            new_column_id = f"c_{column_last_index}"
            
            # Create new row based on artifact_side
            new_row = {
                'stage_id': stage_id,
                'stage_name': stage_name,
                'artifact_id': artifact_id,
                'artifact_name': artifact_name,
                'column_id': new_column_id,
                'column_name': target_column_name
            }
            
            if artifact_side == 'source':
                # For source side: use column_name as-is
                new_row['data_type'] = upstream_data_type
                new_row['order'] = upstream_order
                new_row['column_business_name'] = upstream_column_business_name
                new_row['column_group'] = upstream_column_group
                new_row['column_comment'] = upstream_column_comment
                
            elif artifact_side == 'business':
                # Convert data type based on platform mapping
                converted_data_type = self._convert_data_type(
                    upstream_data_type, 
                    platform, 
                    df_data_mappings
                )
                new_row['data_type'] = converted_data_type
                
                # FIX 2: If column_group is 'attribute', start order at 100 and increment by 1
                if pd.notna(upstream_column_group) and str(upstream_column_group).lower() == 'attribute':
                    new_row['order'] = attribute_order_counter
                    attribute_order_counter += 1
                else:
                    # For non-attributes, use original logic
                    new_row['order'] = upstream_order + 100
                
                # Clear business fields
                new_row['column_business_name'] = ''
                new_row['column_group'] = ''
                new_row['column_comment'] = ''
            
            else:
                # Default behavior if artifact_side is not recognized
                self.logger.warning(f"    Unknown artifact_side: {artifact_side}, using source logic")
                new_row['data_type'] = upstream_data_type
                new_row['order'] = upstream_order
                new_row['column_business_name'] = upstream_column_business_name
                new_row['column_group'] = upstream_column_group
                new_row['column_comment'] = upstream_column_comment
            
            new_rows.append(new_row)
        
        # Append new rows to the columns sheet
        if new_rows:
            self.logger.info(f"    Inserting {len(new_rows)} new fields for artifact {artifact_id}")
            df_new_rows = pd.DataFrame(new_rows)
            
            # Append to Excel file (preserve structure, don't overwrite headers)
            self.excel_utils.append_data_preserve_structure(
                self.workbook_path, 
                'columns', 
                df_new_rows
            )
            
            self.logger.info(f"    Successfully cascaded {len(new_rows)} fields from {upstream_art_id} to {artifact_id}")
    
    def _convert_data_type(self, source_data_type: str, platform: str, 
                          df_data_mappings: pd.DataFrame) -> str:
        """
        Convert data type from source to target platform using mapping table.
        
        Args:
            source_data_type: Original data type
            platform: Target platform name
            df_data_mappings: DataFrame with data type mappings
            
        Returns:
            Converted data type for target platform
        """
        try:
            # Normalize source data type for lookup
            source_type_normalized = source_data_type.upper().strip()
            
            # Find the source type in the first column ('source' or similar)
            # Assuming first column is the source column
            first_col = df_data_mappings.columns[0]
            
            # Look up the row with matching source type
            matching_row = df_data_mappings[df_data_mappings[first_col].str.upper() == source_type_normalized]
            
            if matching_row.empty:
                self.logger.warning(f"      Data type '{source_data_type}' not found in mappings, using as-is")
                return source_data_type
            
            # Get the value from the platform column
            if platform in df_data_mappings.columns:
                converted_type = matching_row.iloc[0][platform]
                self.logger.info(f"      Converted {source_data_type} -> {converted_type} for platform {platform}")
                return converted_type
            else:
                self.logger.warning(f"      Platform '{platform}' not found in mappings, using source type")
                return source_data_type
                
        except Exception as e:
            self.logger.warning(f"      Error converting data type: {str(e)}, using source type")
            return source_data_type
    
    def _cleanup_duplicate_columns(self):
        """
        Clean up duplicate column names within each artifact.
        Keeps the first occurrence and removes subsequent duplicates.
        """
        try:
            # Read current columns sheet
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            
            if df_columns.empty:
                self.logger.info("  No columns to clean up")
                return
            
            initial_count = len(df_columns)
            self.logger.info(f"  Initial column count: {initial_count}")
            
            # Group by artifact_id and remove duplicates within each group
            artifacts = df_columns['artifact_id'].unique()
            
            cleaned_rows = []
            total_duplicates_removed = 0
            
            for artifact_id in artifacts:
                artifact_rows = df_columns[df_columns['artifact_id'] == artifact_id]
                
                # Find duplicates based on column_name within this artifact
                before_count = len(artifact_rows)
                
                # Keep first occurrence, mark duplicates
                artifact_rows_cleaned = artifact_rows.drop_duplicates(
                    subset=['artifact_id', 'column_name'], 
                    keep='first'
                )
                
                after_count = len(artifact_rows_cleaned)
                duplicates_removed = before_count - after_count
                
                if duplicates_removed > 0:
                    self.logger.info(f"  Artifact {artifact_id}: Removed {duplicates_removed} duplicate column(s)")
                    total_duplicates_removed += duplicates_removed
                
                cleaned_rows.append(artifact_rows_cleaned)
            
            # Combine all cleaned artifacts
            df_cleaned = pd.concat(cleaned_rows, ignore_index=True)
            
            final_count = len(df_cleaned)
            self.logger.info(f"  Final column count: {final_count}")
            self.logger.info(f"  Total duplicates removed: {total_duplicates_removed}")
            
            if total_duplicates_removed > 0:
                # Write cleaned data back to Excel
                self.logger.info("  Writing cleaned data back to workbook...")
                
                # Use openpyxl to replace the columns sheet
                from openpyxl import load_workbook
                
                wb = load_workbook(self.workbook_path)
                
                # Remove old columns sheet
                if 'columns' in wb.sheetnames:
                    del wb['columns']
                
                # Create new columns sheet
                ws = wb.create_sheet('columns', 2)  # Insert at position 2 (after artifacts and stages)
                
                # Write headers
                headers = df_cleaned.columns.tolist()
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Write data
                for row_idx, row_data in enumerate(df_cleaned.values, start=2):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(self.workbook_path)
                self.logger.info("  Cleanup completed successfully")
            else:
                self.logger.info("  No duplicates found, no cleanup needed")
                
        except Exception as e:
            self.logger.error(f"Error during duplicate cleanup: {str(e)}")
            raise


def main():
    """Main function for testing cascade operations."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python a_cascade_operations.py <workbook_path>")
        sys.exit(1)
    
    workbook_path = sys.argv[1]
    
    cascade_ops = CascadeOperations(workbook_path)
    cascade_ops.cascade_main()


if __name__ == "__main__":
    main()
