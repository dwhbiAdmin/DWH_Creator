"""
Cascade Re-enumeration Module
==============================

This module handles re-enumeration and reordering of columns after cascade operations.
It should be run as the final step after all enhancements are complete.
"""

import pandas as pd
from pathlib import Path
import sys

# Add src directory to path for imports
current_dir = Path(__file__).parent
src_dir = current_dir.parent.parent.parent
sys.path.insert(0, str(src_dir))

from utils.c_workbench_excel_utils import ExcelUtils
from utils.z_logger import Logger


class CascadeReenumeration:
    """Handles re-enumeration and reordering of columns."""
    
    def __init__(self, workbook_path: str):
        """
        Initialize the Cascade Re-enumeration.
        
        Args:
            workbook_path: Path to the workbook Excel file
        """
        self.workbook_path = workbook_path
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
    
    def run_all_reenumeration(self):
        """
        Run all re-enumeration and reordering operations.
        
        This is the main method that should be called as the final step.
        It runs:
        1. Reorder columns within each artifact
        2. Re-enumerate column IDs sequentially
        """
        try:
            self.logger.info("Starting cascade re-enumeration...")
            
            # Step 1: Reorder columns within artifacts
            self.reorder_columns_within_artifacts()
            
            # Step 2: Re-enumerate column IDs
            self.reenumerate_column_ids()
            
            self.logger.info("All re-enumeration operations completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error in cascade re-enumeration: {str(e)}")
            raise
    
    def reorder_columns_within_artifacts(self):
        """
        Reorder columns within each artifact based on the 'order' field.
        This ensures columns are in the correct sequence within each artifact.
        """
        try:
            self.logger.info("Step 1: Reordering columns within artifacts...")
            
            # Read current columns sheet
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            
            if df_columns.empty:
                self.logger.info("  No columns to reorder")
                return
            
            # Sort by artifact_id and order
            df_sorted = df_columns.sort_values(by=['artifact_id', 'order'], ascending=[True, True])
            
            # Check if order changed
            if not df_columns.equals(df_sorted):
                self.logger.info(f"  Columns reordered within {df_columns['artifact_id'].nunique()} artifact(s)")
                
                # Write sorted data back to Excel
                self._write_columns_to_excel(df_sorted)
                self.logger.info("  Reordering completed successfully")
            else:
                self.logger.info("  Columns already in correct order")
                
        except Exception as e:
            self.logger.error(f"Error during column reordering: {str(e)}")
            raise
    
    def reenumerate_column_ids(self):
        """
        Re-enumerate column IDs to be sequential (c_1, c_2, c_3, etc.).
        This should be run at the end of all enhancements.
        """
        try:
            self.logger.info("Step 2: Re-enumerating column IDs...")
            
            # Read current columns sheet
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            
            if df_columns.empty:
                self.logger.info("  No columns to re-enumerate")
                return
            
            initial_first_id = df_columns.iloc[0]['column_id'] if len(df_columns) > 0 else None
            initial_last_id = df_columns.iloc[-1]['column_id'] if len(df_columns) > 0 else None
            
            self.logger.info(f"  Current column IDs: {initial_first_id} to {initial_last_id}")
            
            # Re-enumerate column IDs sequentially
            for idx, row_idx in enumerate(df_columns.index, start=1):
                df_columns.at[row_idx, 'column_id'] = f"c_{idx}"
            
            final_first_id = df_columns.iloc[0]['column_id']
            final_last_id = df_columns.iloc[-1]['column_id']
            
            self.logger.info(f"  Re-enumerated column IDs: {final_first_id} to {final_last_id}")
            self.logger.info(f"  Total columns: {len(df_columns)}")
            
            # Write re-enumerated data back to Excel
            self._write_columns_to_excel(df_columns)
            self.logger.info("  Re-enumeration completed successfully")
            
        except Exception as e:
            self.logger.error(f"Error during column ID re-enumeration: {str(e)}")
            raise
    
    def _write_columns_to_excel(self, df_columns: pd.DataFrame):
        """
        Write columns DataFrame back to Excel, replacing the columns sheet.
        
        Args:
            df_columns: DataFrame containing the columns data
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.workbook_path)
            
            # Remove old columns sheet
            if 'columns' in wb.sheetnames:
                del wb['columns']
            
            # Create new columns sheet at position 2 (after stages and artifacts)
            ws = wb.create_sheet('columns', 2)
            
            # Write headers
            headers = df_columns.columns.tolist()
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(df_columns.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(self.workbook_path)
            
        except Exception as e:
            self.logger.error(f"Error writing to Excel: {str(e)}")
            raise


def main():
    """Main function for testing cascade re-enumeration."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python c_cascade_reenumeration.py <workbook_path>")
        sys.exit(1)
    
    workbook_path = sys.argv[1]
    
    reenumeration = CascadeReenumeration(workbook_path)
    reenumeration.run_all_reenumeration()


if __name__ == "__main__":
    main()
