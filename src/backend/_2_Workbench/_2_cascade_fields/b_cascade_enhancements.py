"""
Cascade Enhancements Module
============================

This module provides enhancement operations to be run after cascade operations.
It includes adding technical columns with special roles and cleanup of duplicates.
Re-enumeration is handled in c_cascade_reenumeration module.
"""

import pandas as pd
from pathlib import Path
import sys

# Add src directory to path for imports
current_dir = Path(__file__).parent
src_dir = current_dir.parent.parent.parent
sys.path.insert(0, str(src_dir))

from utils.c_workbench_excel_utils import ExcelUtils
from utils.z_logger import Logger
from backend._2_Workbench._2_cascade_fields.c_cascade_reenumeration import CascadeReenumeration


class CascadeEnhancements:
    """Handles post-cascade enhancement operations."""
    
    def __init__(self, workbook_path: str):
        """
        Initialize the Cascade Enhancements.
        
        Args:
            workbook_path: Path to the workbook Excel file
        """
        self.workbook_path = workbook_path
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
    
    def run_all_enhancements(self):
        """
        Run all enhancement operations in sequence.
        
        This is the main method that should be called after cascade operations.
        It runs:
        1. Add technical columns with special role handling
        2. Cleanup duplicate columns
        3. Re-enumeration (via c_cascade_reenumeration module)
        """
        try:
            self.logger.info("Starting cascade enhancements...")
            
            # Step 1: Add technical columns with special role handling
            self.add_technical_columns_with_special_roles()
            
            # Step 2: Clean up duplicate columns
            self.cleanup_duplicate_columns()
            
            # Step 3: Run re-enumeration (in separate module)
            self.logger.info("Running re-enumeration module...")
            reenumeration = CascadeReenumeration(self.workbook_path)
            reenumeration.run_all_reenumeration()
            
            self.logger.info("All cascade enhancements completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error in cascade enhancements: {str(e)}")
            raise
    
    def add_technical_columns_with_special_roles(self):
        """
        Add technical columns from conf_2_technical_columns with special role handling.
        
        Logic:
        1. Iterate through artifacts
        2. For each artifact, read technical fields for its stage_id from conf_2_technical_columns
        3. Add fields without special roles ({}) directly
        4. Handle special roles:
           - stage_id=s2: Add {field_name}_PK for each primary key from s1 upstream
           - {table_name} in dimension: Replace with current artifact name
           - {table_name} in fact: Add {table_name}_SK and {table_name}_BK for each upstream dimension
        """
        try:
            self.logger.info("Adding technical columns with special role handling...")
            
            # Read required sheets
            df_artifacts = pd.read_excel(self.workbook_path, sheet_name='artifacts')
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            df_tech_columns = pd.read_excel(self.workbook_path, sheet_name='conf_2_technical_columns')
            
            # Track new columns to add
            new_columns = []
            
            # Iterate through each artifact
            for _, artifact in df_artifacts.iterrows():
                artifact_id = artifact['artifact_id']
                artifact_name = artifact['artifact_name']
                artifact_type = artifact.get('artifact_type', '')
                stage_id = artifact.get('stage_id', '')
                upstream_artifact = artifact.get('upstream_artifact', '')
                
                self.logger.info(f"Processing artifact {artifact_id} ({artifact_name}), stage={stage_id}, type={artifact_type}")
                
                # Get technical fields for this stage
                df_tech_fields = df_tech_columns[df_tech_columns['stage_id'] == stage_id]
                
                if df_tech_fields.empty:
                    self.logger.info(f"  No technical fields defined for stage {stage_id}")
                    continue
                
                # Get current max order for this artifact
                artifact_columns = df_columns[df_columns['artifact_id'] == artifact_id]
                max_order = artifact_columns['order'].max() if not artifact_columns.empty else 0
                if pd.isna(max_order):
                    max_order = 0
                
                # Process each technical field
                for _, tech_field in df_tech_fields.iterrows():
                    column_name = tech_field['column_name']
                    data_type = tech_field['data_type']
                    group = tech_field.get('group', 'technical')
                    order = tech_field.get('order', 0)
                    
                    # Check if this field contains special roles (placeholders in {})
                    has_special_role = '{' in str(column_name) and '}' in str(column_name)
                    
                    if not has_special_role:
                        # Simple field - add directly
                        new_col = self._create_column_entry(
                            artifact_id=artifact_id,
                            column_name=column_name,
                            data_type=data_type,
                            group=group,
                            order=max_order + order
                        )
                        new_columns.append(new_col)
                        self.logger.info(f"  Added technical field: {column_name}")
                    
                    else:
                        # Handle special roles
                        self._handle_special_role_field(
                            column_name=column_name,
                            data_type=data_type,
                            group=group,
                            base_order=max_order + order,
                            artifact_id=artifact_id,
                            artifact_name=artifact_name,
                            artifact_type=artifact_type,
                            stage_id=stage_id,
                            upstream_artifact=upstream_artifact,
                            df_columns=df_columns,
                            df_artifacts=df_artifacts,
                            new_columns=new_columns
                        )
            
            # Add all new columns to the sheet
            if new_columns:
                self.logger.info(f"Adding {len(new_columns)} technical column(s) to columns sheet...")
                df_new = pd.DataFrame(new_columns)
                df_combined = pd.concat([df_columns, df_new], ignore_index=True)
                self._write_columns_to_excel(df_combined)
                self.logger.info("Technical columns added successfully")
            else:
                self.logger.info("No technical columns to add")
                
        except Exception as e:
            self.logger.error(f"Error adding technical columns: {str(e)}")
            raise
    
    def _handle_special_role_field(self, column_name: str, data_type: str, group: str, base_order: int,
                                   artifact_id: str, artifact_name: str, artifact_type: str, stage_id: str,
                                   upstream_artifact: str, df_columns: pd.DataFrame, df_artifacts: pd.DataFrame,
                                   new_columns: list):
        """
        Handle special role fields with placeholders like {field_name}_PK or {table_name}_SK.
        
        Args:
            column_name: Template column name with placeholders
            data_type: Data type (may contain placeholders)
            group: Column group
            base_order: Base order number
            artifact_id: Current artifact ID
            artifact_name: Current artifact name
            artifact_type: Artifact type (dimension, fact, etc.)
            stage_id: Current stage ID
            upstream_artifact: Upstream artifact(s)
            df_columns: DataFrame of existing columns
            df_artifacts: DataFrame of all artifacts
            new_columns: List to append new column entries
        """
        # Special role 1: stage_id=s2, add {field_name}_PK for each primary key from s1 upstream
        if stage_id == 's2' and '{field_name}_PK' in column_name:
            self._add_primary_key_fields_from_upstream(
                artifact_id, upstream_artifact, df_columns, df_artifacts, 
                data_type, group, base_order, new_columns
            )
        
        # Special role 2: {table_name} in dimension - replace with current artifact name
        elif '{table_name}' in column_name and artifact_type == 'dimension':
            resolved_column_name = column_name.replace('{table_name}', artifact_name)
            new_col = self._create_column_entry(
                artifact_id=artifact_id,
                column_name=resolved_column_name,
                data_type=data_type,
                group=group,
                order=base_order
            )
            new_columns.append(new_col)
            self.logger.info(f"  Added dimension field: {resolved_column_name}")
        
        # Special role 3: {table_name}_SK/_BK in fact - add for each upstream dimension
        elif '{table_name}' in column_name and artifact_type == 'fact':
            self._add_sk_bk_fields_for_fact(
                artifact_id, upstream_artifact, column_name, df_columns, df_artifacts,
                group, base_order, new_columns
            )
        
        else:
            self.logger.warning(f"  Unhandled special role field: {column_name}")
    
    def _add_primary_key_fields_from_upstream(self, artifact_id: str, upstream_artifact: str,
                                              df_columns: pd.DataFrame, df_artifacts: pd.DataFrame,
                                              data_type_template: str, group: str, base_order: int,
                                              new_columns: list):
        """Add {field_name}_PK fields for each primary key from s1 upstream artifacts."""
        if not upstream_artifact:
            self.logger.info(f"  No upstream artifact for {artifact_id}, skipping PK fields")
            return
        
        # Split multiple upstream artifacts
        upstream_ids = [u.strip() for u in upstream_artifact.split(';') if u.strip()]
        
        for upstream_id in upstream_ids:
            # Get upstream artifact stage
            upstream_row = df_artifacts[df_artifacts['artifact_id'] == upstream_id]
            if upstream_row.empty:
                self.logger.warning(f"  Upstream artifact {upstream_id} not found")
                continue
            
            upstream_stage = upstream_row.iloc[0].get('stage_id', '')
            
            if upstream_stage == 's1':
                # Get primary keys from upstream s1 artifact
                upstream_columns = df_columns[
                    (df_columns['artifact_id'] == upstream_id) & 
                    (df_columns['column_group'] == 'primary key')
                ]
                
                for idx, pk_col in upstream_columns.iterrows():
                    pk_name = pk_col['column_name']
                    pk_data_type = pk_col['data_type']
                    
                    # Create {field_name}_PK
                    new_column_name = f"{pk_name}_PK"
                    
                    new_col = self._create_column_entry(
                        artifact_id=artifact_id,
                        column_name=new_column_name,
                        data_type=pk_data_type,
                        group=group,
                        order=base_order
                    )
                    new_columns.append(new_col)
                    self.logger.info(f"  Added PK field from s1: {new_column_name}")
                    base_order += 1
    
    def _add_sk_bk_fields_for_fact(self, artifact_id: str, upstream_artifact: str, column_template: str,
                                   df_columns: pd.DataFrame, df_artifacts: pd.DataFrame,
                                   group: str, base_order: int, new_columns: list):
        """Add {table_name}_SK and {table_name}_BK fields for each upstream dimension."""
        if not upstream_artifact:
            self.logger.info(f"  No upstream artifact for fact {artifact_id}, skipping SK/BK fields")
            return
        
        # Split multiple upstream artifacts
        upstream_ids = [u.strip() for u in upstream_artifact.split(';') if u.strip()]
        
        for upstream_id in upstream_ids:
            # Get upstream artifact info
            upstream_row = df_artifacts[df_artifacts['artifact_id'] == upstream_id]
            if upstream_row.empty:
                self.logger.warning(f"  Upstream artifact {upstream_id} not found")
                continue
            
            upstream_type = upstream_row.iloc[0].get('artifact_type', '')
            upstream_name = upstream_row.iloc[0].get('artifact_name', '')
            
            if upstream_type == 'dimension':
                # Get business key from upstream dimension to determine data type
                upstream_bk = df_columns[
                    (df_columns['artifact_id'] == upstream_id) & 
                    (df_columns['column_group'] == 'business key')
                ]
                
                bk_data_type = 'STRING'  # default
                if not upstream_bk.empty:
                    bk_data_type = upstream_bk.iloc[0]['data_type']
                
                # Add {table_name}_SK
                if '_SK' in column_template:
                    sk_column_name = f"{upstream_name}_SK"
                    new_col = self._create_column_entry(
                        artifact_id=artifact_id,
                        column_name=sk_column_name,
                        data_type='INT',  # Surrogate keys are typically INT
                        group=group,
                        order=base_order
                    )
                    new_columns.append(new_col)
                    self.logger.info(f"  Added SK field: {sk_column_name}")
                    base_order += 1
                
                # Add {table_name}_BK
                if '_BK' in column_template:
                    bk_column_name = f"{upstream_name}_BK"
                    new_col = self._create_column_entry(
                        artifact_id=artifact_id,
                        column_name=bk_column_name,
                        data_type=bk_data_type,
                        group=group,
                        order=base_order
                    )
                    new_columns.append(new_col)
                    self.logger.info(f"  Added BK field: {bk_column_name}")
                    base_order += 1
    
    def _create_column_entry(self, artifact_id: str, column_name: str, data_type: str, 
                           group: str, order: int) -> dict:
        """Create a new column entry dictionary with new fields structure."""
        return {
            'column_id': f"c_temp_{len(str(order))}",  # Temporary ID, will be re-enumerated
            'artifact_id': artifact_id,
            'column_name': column_name,
            'data_type': data_type,
            'column_group': group,
            'order': order,
            'column_business_name': column_name.lower(),
            'column_comment': f'Technical column: {column_name}',
            'cascaded': True,
            # New fields (columns I-M)
            'source_column_name': column_name,  # Default to same as column_name
            'lookup_fields': '',  # Empty for technical columns
            'etl_simple_trnasformation': '',  # Empty by default
            'ai_transformation_prompt': '',  # Empty by default
            'etl_ai_transformation': ''  # Empty by default
        }
    
    def cleanup_duplicate_columns(self):
        """
        Clean up duplicate column names within each artifact.
        Keeps the first occurrence and removes subsequent duplicates.
        """
        try:
            self.logger.info("Step 1: Cleaning up duplicate columns...")
            
            # Read current columns sheet
            df_columns = pd.read_excel(self.workbook_path, sheet_name='columns')
            
            if df_columns.empty:
                self.logger.info("  No columns to clean up")
                return
            
            initial_count = len(df_columns)
            self.logger.info(f"  Initial column count: {initial_count}")
            
            # Group by artifact_id and remove duplicates within each group
            artifacts = df_columns['artifact_id'].unique()
            
            cleaned_rows = []
            total_duplicates_removed = 0
            
            for artifact_id in artifacts:
                artifact_rows = df_columns[df_columns['artifact_id'] == artifact_id]
                
                # Find duplicates based on column_name within this artifact
                before_count = len(artifact_rows)
                
                # Keep first occurrence, remove duplicates
                artifact_rows_cleaned = artifact_rows.drop_duplicates(
                    subset=['artifact_id', 'column_name'], 
                    keep='first'
                )
                
                after_count = len(artifact_rows_cleaned)
                duplicates_removed = before_count - after_count
                
                if duplicates_removed > 0:
                    self.logger.info(f"  Artifact {artifact_id}: Removed {duplicates_removed} duplicate column(s)")
                    total_duplicates_removed += duplicates_removed
                
                cleaned_rows.append(artifact_rows_cleaned)
            
            # Combine all cleaned artifacts
            df_cleaned = pd.concat(cleaned_rows, ignore_index=True)
            
            final_count = len(df_cleaned)
            self.logger.info(f"  Final column count: {final_count}")
            self.logger.info(f"  Total duplicates removed: {total_duplicates_removed}")
            
            if total_duplicates_removed > 0:
                # Write cleaned data back to Excel
                self._write_columns_to_excel(df_cleaned)
                self.logger.info("  Cleanup completed successfully")
            else:
                self.logger.info("  No duplicates found, no cleanup needed")
                
        except Exception as e:
            self.logger.error(f"Error during duplicate cleanup: {str(e)}")
            raise
    
    def _write_columns_to_excel(self, df_columns: pd.DataFrame):
        """
        Write columns DataFrame back to Excel, replacing the columns sheet.
        
        Args:
            df_columns: DataFrame containing the columns data
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.workbook_path)
            
            # Remove old columns sheet
            if 'columns' in wb.sheetnames:
                del wb['columns']
            
            # Create new columns sheet at position 2 (after stages and artifacts)
            ws = wb.create_sheet('columns', 2)
            
            # Write headers
            headers = df_columns.columns.tolist()
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(df_columns.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(self.workbook_path)
            
        except Exception as e:
            self.logger.error(f"Error writing to Excel: {str(e)}")
            raise


def main():
    """Main function for testing cascade enhancements."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python b_cascade_enhancements.py <workbook_path>")
        sys.exit(1)
    
    workbook_path = sys.argv[1]
    
    enhancements = CascadeEnhancements(workbook_path)
    enhancements.run_all_enhancements()


if __name__ == "__main__":
    main()
