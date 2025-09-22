"""
Workbench Manager - Step 2
==========================

Handles Excel workbook operations for stages, artifacts, and columns sheets.
Implements Tab B functionality from the specification.

Key responsibilities:
- Managing the three main sheets: Stages, Artifacts, Columns
- Import/Assign operations from source files
- AI comment generation for artifacts and columns
- Cascade operations for deterministic column filling
- Sync and validation operations
"""

# ANCHOR: Imports and Dependencies

import os
import subprocess
import pandas as pd
import glob
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# Import utilities with proper path handling
import sys
current_dir = Path(__file__).parent
src_dir = current_dir.parent
sys.path.insert(0, str(src_dir))

from utils.B_excel_utils import ExcelUtils
from utils.logger import Logger
from utils.B_worksheet_config_manager import ConfigManager
from utils.B_workbench_cascading import ColumnCascadingEngine
from .Y_ai_manager import AIWorkbenchManager

# ANCHOR: WorkbenchManager Class Definition

class WorkbenchManager:
    """
    Manages Excel workbook operations and sheet manipulations.
    """
    
    # ANCHOR: Initialization and Setup Methods
    def __init__(self, project_path: str = None, openai_api_key: str = None):
        """
        Initialize the Workbench Manager.
        
        Args:
            project_path: Path to the project directory
            openai_api_key: OpenAI API key for AI comment generation
        """
        self.project_path = project_path
        self.workbook_path = None
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
        self.config = ConfigManager()
        self.openai_api_key = openai_api_key
        self.ai_workbench = None  # Will be initialized when workbook is found
        self.cascading_engine = None  # Will be initialized when workbook is found
        
        if project_path:
            self._find_workbook()
    
    def _find_workbook(self):
        """Find the workbook file in the project."""
        if not self.project_path:
            return
            
        workbench_dir = os.path.join(self.project_path, "2_workbench")
        workbook_pattern = os.path.join(workbench_dir, "workbench_*.xlsx")
        workbooks = glob.glob(workbook_pattern)
        
        # Filter out temporary files and configuration files
        workbooks = [wb for wb in workbooks if not os.path.basename(wb).startswith("~$") and "configuration" not in os.path.basename(wb)]
        
        if workbooks:
            self.workbook_path = workbooks[0]
            self.logger.info(f"Found workbook: {self.workbook_path}")
            # Initialize AI workbench manager
            self.ai_workbench = AIWorkbenchManager(self.workbook_path, self.openai_api_key)
            # Initialize column cascading engine
            # Initialize cascading engine with workbench manager reference
            self.cascading_engine = ColumnCascadingEngine(self.workbook_path)
            self.cascading_engine.workbench_manager = self  # Pass reference for AI access
        else:
            self.logger.error(f"No workbook found in {workbench_dir}")

    # ANCHOR: Sheet Opening Methods
    def open_stages_sheet(self) -> bool:
        """Open the Stages sheet for editing."""
        if not self.workbook_path:
            self.logger.error("No workbook path available")
            return False
            
        try:
            if os.name == 'nt':  # Windows
                os.startfile(self.workbook_path)
            elif os.name == 'posix':  # macOS and Linux
                subprocess.call(['open', self.workbook_path])
                
            self.logger.info("Opened Excel workbook for Stages sheet editing")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to open workbook: {str(e)}")
            return False
    
    def open_artifacts_sheet(self) -> bool:
        """Open the Artifacts sheet for editing."""
        return self.open_stages_sheet()  # Same operation - opens Excel
    
    def open_columns_sheet(self) -> bool:
        """Open the Columns sheet for editing."""
        return self.open_stages_sheet()  # Same operation - opens Excel

    # ANCHOR: Import and CSV Processing Methods
    def import_assign_columns(self, source_folder: str = None) -> bool:
        """
        Import and assign columns from source files.
        
        Args:
            source_folder: Path to 1_sources folder (defaults to project's 1_sources)
            
        Returns:
            bool: True if import successful
        """
        try:
            # Determine source folder
            if source_folder is None:
                source_folder = os.path.join(self.project_path, "1_data_sources")
            
            if not os.path.exists(source_folder):
                self.logger.error(f"Source folder not found: {source_folder}")
                return False
            
            # Get CSV files from source folder
            csv_files = self._get_csv_files(source_folder)
            if not csv_files:
                self.logger.info(f"No CSV files found in {source_folder}")
                return True
            
            self.logger.info(f"Found {len(csv_files)} CSV files to process")
            
            # Load workbook data
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "artifacts")
            if artifacts_df.empty:
                self.logger.error("No artifacts found in workbook")
                return False
            
            total_columns_added = 0
            processed_files = []
            failed_files = []
            
            # Process each CSV file
            for csv_file in csv_files:
                try:
                    csv_filename = os.path.basename(csv_file)
                    self.logger.info(f"Processing: {csv_filename}")
                    
                    # Analyze CSV structure
                    columns_info = self._analyze_csv_file(csv_file)
                    if not columns_info:
                        failed_files.append(csv_filename)
                        continue
                    
                    # Find corresponding artifact
                    artifact_id = self._find_artifact_id(csv_filename, artifacts_df)
                    if not artifact_id:
                        self.logger.warning(f"No artifact found for {csv_filename}")
                        failed_files.append(csv_filename)
                        continue
                    
                    # Update columns sheet
                    success = self._add_columns_to_workbook(artifact_id, columns_info)
                    if success:
                        total_columns_added += len(columns_info)
                        processed_files.append(csv_filename)
                        self.logger.info(f"Added {len(columns_info)} columns for {csv_filename}")
                    else:
                        failed_files.append(csv_filename)
                        self.logger.error(f"Failed to add columns for {csv_filename}")
                    
                except Exception as e:
                    self.logger.error(f"Error processing {csv_file}: {str(e)}")
                    failed_files.append(os.path.basename(csv_file))
                    continue
            
            # Report results
            if failed_files:
                self.logger.error(f"Import completed with errors: {len(processed_files)} files successful, {len(failed_files)} files failed")
                self.logger.error(f"Failed files: {', '.join(failed_files)}")
                return len(processed_files) > 0  # Return True if at least some files were processed
            else:
                self.logger.info(f"Import completed successfully: {len(processed_files)} files, {total_columns_added} columns")
                return True
            
        except Exception as e:
            self.logger.error(f"Failed to import/assign columns: {str(e)}")
            return False

    # ANCHOR: CSV Analysis Helper Methods
    def _get_csv_files(self, source_folder: str) -> List[str]:
        """Get all CSV files from source folder."""
        csv_pattern = os.path.join(source_folder, "*.csv")
        return glob.glob(csv_pattern)
    
    def _analyze_csv_file(self, csv_path: str) -> List[Dict]:
        """Analyze CSV file and extract column metadata."""
        try:
            # Try different separators
            try:
                df = pd.read_csv(csv_path, sep=';', nrows=100)  # Sample first 100 rows
                if len(df.columns) == 1:
                    df = pd.read_csv(csv_path, sep=',', nrows=100)
            except Exception:
                df = pd.read_csv(csv_path, sep=',', nrows=100)
            
            columns_info = []
            for idx, column in enumerate(df.columns):
                data_type = self._detect_data_type(df[column])
                columns_info.append({
                    'column_name': column,
                    'data_type': data_type,
                    'order': idx + 1
                })
            
            return columns_info
            
        except Exception as e:
            self.logger.error(f"Error analyzing CSV {csv_path}: {str(e)}")
            return []
    
    def _detect_data_type(self, series: pd.Series) -> str:
        """Detect data type of a pandas Series."""
        if pd.api.types.is_integer_dtype(series):
            return 'integer'
        elif pd.api.types.is_numeric_dtype(series):
            return 'decimal'
        elif pd.api.types.is_datetime64_any_dtype(series):
            return 'datetime'
        elif pd.api.types.is_bool_dtype(series):
            return 'boolean'
        else:
            # For object type, try to infer
            try:
                pd.to_numeric(series.dropna())
                return 'decimal'
            except:
                try:
                    pd.to_datetime(series.dropna())
                    return 'datetime'
                except:
                    return 'string'
    
    def _find_artifact_id(self, csv_filename: str, artifacts_df: pd.DataFrame) -> str:
        """Find artifact ID for a CSV filename."""
        # Look for exact match in Artifact Name column
        matches = artifacts_df[artifacts_df['artifact_name'] == csv_filename]
        if not matches.empty:
            return matches.iloc[0]['artifact_id']
        
        # Try without extension
        filename_without_ext = os.path.splitext(csv_filename)[0]
        matches = artifacts_df[artifacts_df['artifact_name'] == filename_without_ext]
        if not matches.empty:
            return matches.iloc[0]['artifact_id']
        
        return None

    # ANCHOR: Workbook Data Management Methods
    def _add_columns_to_workbook(self, artifact_id: str, columns_info: List[Dict]) -> bool:
        """Add column information to the columns sheet."""
        try:
            # Read current columns data
            columns_df = self.excel_utils.read_sheet_data(self.workbook_path, "columns")
            
            # Remove existing columns for this artifact
            if not columns_df.empty:
                columns_df = columns_df[columns_df['artifact_id'] != artifact_id]
            
            # Get artifact information for additional fields
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "artifacts")
            artifact_row = artifacts_df[artifacts_df['artifact_id'] == artifact_id]
            if artifact_row.empty:
                self.logger.error(f"Artifact {artifact_id} not found in artifacts sheet")
                return False
            
            artifact_info = artifact_row.iloc[0]
            stage_name = artifact_info.get('stage_name', '')
            artifact_name = artifact_info.get('artifact_name', '')
            
            # Map stage name to stage_id
            stage_id_mapping = {
                '0_drop_zone': 's0',
                '1_bronze': 's1', 
                '2_silver': 's2',
                '3_gold': 's3',
                '4_mart': 's4',
                '5_PBI_Model': 's5',
                '6_PBI_Reports': 's6'
            }
            stage_id = stage_id_mapping.get(stage_name, 's0')
            
            # Create new rows for this artifact using proper column structure
            new_rows = []
            for col_info in columns_info:
                # Use the established protected column structure from session context
                # IMPORTANT: Match exact column order from workbench setup
                new_row = {
                    'stage_id': stage_id,
                    'stage_name': stage_name,
                    'artifact_id': artifact_id,
                    'artifact_name': artifact_name,
                    'column_id': f"c{col_info['order']}",
                    'column_name': col_info['column_name'],
                    'data_type': col_info['data_type'],  # Correct position
                    'order': col_info['order'],           # Correct position
                    'column_business_name': '',  # Will be filled by AI
                    'column_group': '',  # Will be filled by cascade (lowercase as per session context)
                    'column_comment': ''  # Will be filled by AI
                }
                new_rows.append(new_row)
            
            # Remove existing columns for this artifact from Excel sheet directly
            self._remove_artifact_columns_from_sheet(artifact_id)
            
            # Append new rows using structure-preserving method
            new_df = pd.DataFrame(new_rows)
            
            # Use append method that preserves headers and formatting
            write_success = self.excel_utils.append_data_preserve_structure(self.workbook_path, "columns", new_df)
            if not write_success:
                self.logger.error(f"Failed to write columns for artifact {artifact_id} to Excel workbook")
                return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding columns to workbook: {str(e)}")
            return False

    def _remove_artifact_columns_from_sheet(self, artifact_id: str) -> bool:
        """Remove existing columns for a specific artifact from the Excel sheet."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.workbook_path)
            if "columns" not in wb.sheetnames:
                return True  # No columns sheet, nothing to remove
            
            ws = wb["columns"]
            
            # Find artifact_id column index (should be column 3 based on structure)
            artifact_id_col = 3
            
            # Find and remove rows with matching artifact_id (start from bottom to avoid index issues)
            rows_to_delete = []
            for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                cell_value = ws.cell(row, artifact_id_col).value
                if cell_value == artifact_id:
                    rows_to_delete.append(row)
            
            # Delete rows from bottom to top to avoid index shifting
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)
            
            # Save workbook
            wb.save(self.workbook_path)
            return True
            
        except Exception as e:
            self.logger.error(f"Error removing artifact columns from sheet: {str(e)}")
            return False

    # ANCHOR: AI Comment Generation Methods
    def generate_ai_comments(self) -> bool:
        """Generate AI comments for artifacts and columns."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return False
        
        return self.ai_workbench.generate_all_ai_comments()

    def generate_artifact_ai_comments(self) -> bool:
        """Generate AI comments for artifacts only."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return False
        
        return self.ai_workbench.generate_artifact_comments()

    def generate_column_ai_comments(self) -> bool:
        """Generate AI comments for columns only."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return False
        
        return self.ai_workbench.generate_column_comments()

    def generate_readable_column_names(self) -> bool:
        """Generate human-readable column names for all columns."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return False
        
        return self.ai_workbench.generate_readable_column_names()

    def get_ai_comment_statistics(self) -> dict:
        """Get AI comment coverage statistics."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return {}
        
        return self.ai_workbench.get_ai_comment_statistics()

    def validate_ai_comments(self) -> dict:
        """Validate AI comment quality."""
        if not self.ai_workbench:
            self.logger.error("AI workbench manager not initialized")
            return {}
        
        return self.ai_workbench.validate_ai_comments()

    def is_ai_available(self) -> bool:
        """Check if AI functionality is available."""
        return self.ai_workbench and self.ai_workbench.is_ai_available()

    # ANCHOR: Cascade and Sync Operations
    def cascade_columns(self, artifact_id: int = None, include_technical_fields: bool = True) -> bool:
        """
        Perform deterministic column filling based on relations.
        
        Args:
            artifact_id: Optional specific artifact ID to cascade. If None, process all artifacts.
            include_technical_fields: Whether to include technical/audit columns
            
        Returns:
            bool: True if cascading was successful
        """
        if not self.cascading_engine:
            self.logger.error("Column cascading engine not initialized")
            return False
        
        try:
            if artifact_id:
                # Cascade for specific artifact
                return self.cascading_engine.cascade_columns_for_artifact(str(artifact_id), include_technical_fields)
            else:
                # Use the new correct cascading logic: find missing artifacts and cascade them
                return self.cascading_engine.cascade_all_missing_artifacts(include_technical_fields)
                
        except Exception as e:
            self.logger.error(f"Error during column cascading: {str(e)}")
            return False
    
    def cascade_columns_for_artifact(self, artifact_id, include_technical_fields: bool = True) -> bool:
        """
        Cascade columns for a specific artifact.
        
        Args:
            artifact_id: ID of the artifact to cascade columns for (string or int)
            include_technical_fields: Whether to include technical/audit columns
            
        Returns:
            bool: True if cascading was successful
        """
        if not self.cascading_engine:
            self.logger.error("Column cascading engine not initialized")
            return False
        
        return self.cascading_engine.cascade_columns_for_artifact(str(artifact_id), include_technical_fields)
    
    def create_cascading_config(self) -> bool:
        """
        Create cascading configuration file with default settings.
        
        Returns:
            bool: True if configuration was created successfully
        """
        if not self.cascading_engine:
            self.logger.error("Column cascading engine not initialized")
            return False
        
        return self.cascading_engine.create_cascading_config_file()
    
    def regenerate_all_columns(self, include_technical_fields: bool = True) -> bool:
        """
        Regenerate ALL columns with globally unique IDs and proper ordering.
        
        Args:
            include_technical_fields: Whether to include technical fields
            
        Returns:
            bool: True if regeneration was successful
        """
        if not self.cascading_engine:
            self.logger.error("Column cascading engine not initialized")
            return False
            
        return self.cascading_engine.regenerate_all_columns_with_unique_ids(include_technical_fields)

    def get_cascading_preview(self, artifact_id: int) -> dict:
        """
        Get a preview of what columns would be cascaded for an artifact.
        
        Args:
            artifact_id: ID of the artifact to preview cascading for
            
        Returns:
            dict: Preview information including column count and names
        """
        if not self.cascading_engine:
            return {"error": "Column cascading engine not initialized"}
        
        try:
            # Load artifacts data
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "Artifacts")
            target_artifact = artifacts_df[artifacts_df['artifact_id'] == artifact_id]
            
            if target_artifact.empty:
                return {"error": f"Artifact {artifact_id} not found"}
            
            target_artifact = target_artifact.iloc[0]
            upstream_relation = target_artifact.get('upstream_relation', '')
            
            if not upstream_relation:
                return {"message": "No upstream relationship defined", "columns": []}
            
            # Get preview (this would be a new method in the cascading engine)
            preview = {
                "artifact_id": artifact_id,
                "artifact_name": target_artifact.get('artifact_name', ''),
                "upstream_relation": upstream_relation,
                "stage": target_artifact.get('stage_name', ''),
                "preview_note": f"Would cascade columns based on '{upstream_relation}' relationship"
            }
            
            return preview
            
        except Exception as e:
            return {"error": f"Error generating preview: {str(e)}"}
    
    def get_artifacts_with_upstream(self) -> List[Dict]:
        """
        Get list of artifacts that have upstream relationships.
        
        Returns:
            List[Dict]: List of artifacts with upstream relationships
        """
        try:
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "Artifacts")
            if artifacts_df.empty:
                return []
            
            # Filter artifacts with upstream relationships
            artifacts_with_upstream = artifacts_df[
                artifacts_df['upstream_relation'].notna() & 
                (artifacts_df['upstream_relation'] != '')
            ]
            
            return artifacts_with_upstream[['artifact_id', 'artifact_name', 'stage_name', 'upstream_relation']].to_dict('records')
            
        except Exception as e:
            self.logger.error(f"Error getting artifacts with upstream relationships: {str(e)}")
            return []
    
    def sync_and_validate(self) -> dict:
        """
        Sync all sheets and run validations.
        
        Returns:
            dict: Validation results and errors
        """
        # TODO: Implement sync and validation logic
        self.logger.info("Sync and validation - implementation pending")
        return {"status": "pending"}

    # ANCHOR: Workbook Management Methods
    def save_workbook(self) -> bool:
        """Save the current workbook version."""
        try:
            # Create backup first
            if self.workbook_path and os.path.exists(self.workbook_path):
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_name = f"workbench_backup_{timestamp}.xlsx"
                backup_path = os.path.join(os.path.dirname(self.workbook_path), backup_name)
                
                import shutil
                shutil.copy2(self.workbook_path, backup_path)
                self.logger.info(f"Created backup: {backup_name}")
                
            self.logger.info("Workbook saved successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {str(e)}")
            return False
