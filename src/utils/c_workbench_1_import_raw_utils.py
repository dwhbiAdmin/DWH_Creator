"""
Raw File Import Utilities
=========================

Dedicated module for handling raw CSV file import functionality.
Extracted and consolidated from c_workbench_manager.py to improve modularity.

Key responsibilities:
- CSV file discovery and analysis
- Data type detection and inference
- Artifact-to-CSV file matching
- Column metadata extraction and workbook integration
- Excel workbook column management for imported data
"""

import os
import glob
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import load_workbook

# Import utilities
import sys
current_dir = Path(__file__).parent
src_dir = current_dir.parent
sys.path.insert(0, str(src_dir))

from utils.c_workbench_excel_utils import ExcelUtils
from utils.z_logger import Logger


class RawFileImporter:
    """
    Handles raw CSV file import operations for workbench integration.
    """
    
    def __init__(self, workbook_path: str = None, project_path: str = None):
        """
        Initialize the Raw File Importer.
        
        Args:
            workbook_path: Path to the Excel workbook
            project_path: Path to the project directory
        """
        self.workbook_path = workbook_path
        self.project_path = project_path
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
    
    # ANCHOR: Main Import Orchestration
    def import_assign_columns(self, source_folder: str = None) -> bool:
        """
        Import and assign columns from source files.
        
        Args:
            source_folder: Path to 1_sources folder (defaults to project's 1_data_sources)
            
        Returns:
            bool: True if import successful
        """
        try:
            # Determine source folder
            if source_folder is None:
                source_folder = os.path.join(self.project_path, "1_data_sources")
            
            if not os.path.exists(source_folder):
                self.logger.error(f"Source folder not found: {source_folder}")
                return False
            
            # Get CSV files from source folder
            csv_files = self.get_csv_files(source_folder)
            if not csv_files:
                self.logger.info(f"No CSV files found in {source_folder}")
                return True
            
            self.logger.info(f"Found {len(csv_files)} CSV files to process")
            
            # Load workbook data
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "artifacts")
            if artifacts_df.empty:
                self.logger.error("No artifacts found in workbook")
                return False
            
            total_columns_added = 0
            processed_files = []
            failed_files = []
            
            # Process each CSV file
            for csv_file in csv_files:
                try:
                    csv_filename = os.path.basename(csv_file)
                    self.logger.info(f"Processing: {csv_filename}")
                    
                    # Analyze CSV structure
                    columns_info = self.analyze_csv_file(csv_file)
                    if not columns_info:
                        failed_files.append(csv_filename)
                        continue
                    
                    # Find corresponding artifact
                    artifact_id = self.find_artifact_id(csv_filename, artifacts_df)
                    if not artifact_id:
                        self.logger.warning(f"No artifact found for {csv_filename}")
                        failed_files.append(csv_filename)
                        continue
                    
                    # Update columns sheet
                    success = self.add_columns_to_workbook(artifact_id, columns_info)
                    if success:
                        total_columns_added += len(columns_info)
                        processed_files.append(csv_filename)
                        self.logger.info(f"Added {len(columns_info)} columns for {csv_filename}")
                    else:
                        failed_files.append(csv_filename)
                        self.logger.error(f"Failed to add columns for {csv_filename}")
                    
                except Exception as e:
                    self.logger.error(f"Error processing {csv_file}: {str(e)}")
                    failed_files.append(os.path.basename(csv_file))
                    continue
            
            # Report results
            if failed_files:
                self.logger.error(f"Import completed with errors: {len(processed_files)} files successful, {len(failed_files)} files failed")
                self.logger.error(f"Failed files: {', '.join(failed_files)}")
                return len(processed_files) > 0  # Return True if at least some files were processed
            else:
                self.logger.info(f"Import completed successfully: {len(processed_files)} files, {total_columns_added} columns")
                return True
            
        except Exception as e:
            self.logger.error(f"Failed to import/assign columns: {str(e)}")
            return False
    
    # ANCHOR: CSV File Discovery and Analysis
    def get_csv_files(self, source_folder: str) -> List[str]:
        """Get all CSV files from source folder."""
        csv_pattern = os.path.join(source_folder, "*.csv")
        return glob.glob(csv_pattern)
    
    def analyze_csv_file(self, csv_path: str) -> List[Dict]:
        """Analyze CSV file and extract column metadata."""
        try:
            # Try different separators
            try:
                df = pd.read_csv(csv_path, sep=';', nrows=100)  # Sample first 100 rows
                if len(df.columns) == 1:
                    df = pd.read_csv(csv_path, sep=',', nrows=100)
            except Exception:
                df = pd.read_csv(csv_path, sep=',', nrows=100)
            
            columns_info = []
            for idx, column in enumerate(df.columns):
                data_type = self.detect_data_type(df[column])
                columns_info.append({
                    'column_name': column,
                    'data_type': data_type,
                    'order': idx + 1
                })
            
            return columns_info
            
        except Exception as e:
            self.logger.error(f"Error analyzing CSV {csv_path}: {str(e)}")
            return []
    
    def detect_data_type(self, series: pd.Series) -> str:
        """Detect data type of a pandas Series."""
        if pd.api.types.is_integer_dtype(series):
            return 'integer'
        elif pd.api.types.is_numeric_dtype(series):
            return 'decimal'
        elif pd.api.types.is_datetime64_any_dtype(series):
            return 'datetime'
        elif pd.api.types.is_bool_dtype(series):
            return 'boolean'
        else:
            # For object type, try to infer
            try:
                pd.to_numeric(series.dropna())
                return 'decimal'
            except:
                try:
                    pd.to_datetime(series.dropna())
                    return 'datetime'
                except:
                    return 'string'
    
    # ANCHOR: Artifact Matching
    def find_artifact_id(self, csv_filename: str, artifacts_df: pd.DataFrame) -> Optional[str]:
        """Find artifact ID for a CSV filename."""
        # Look for exact match in Artifact Name column
        matches = artifacts_df[artifacts_df['artifact_name'] == csv_filename]
        if not matches.empty:
            return matches.iloc[0]['artifact_id']
        
        # Try without extension
        filename_without_ext = os.path.splitext(csv_filename)[0]
        matches = artifacts_df[artifacts_df['artifact_name'] == filename_without_ext]
        if not matches.empty:
            return matches.iloc[0]['artifact_id']
        
        return None
    
    # ANCHOR: Workbook Integration
    def add_columns_to_workbook(self, artifact_id: str, columns_info: List[Dict]) -> bool:
        """Add column information to the columns sheet."""
        try:
            # Read current columns data
            columns_df = self.excel_utils.read_sheet_data(self.workbook_path, "columns")
            
            # Remove existing columns for this artifact
            if not columns_df.empty:
                columns_df = columns_df[columns_df['artifact_id'] != artifact_id]
            
            # Get artifact information for additional fields
            artifacts_df = self.excel_utils.read_sheet_data(self.workbook_path, "artifacts")
            artifact_row = artifacts_df[artifacts_df['artifact_id'] == artifact_id]
            if artifact_row.empty:
                self.logger.error(f"Artifact {artifact_id} not found in artifacts sheet")
                return False
            
            artifact_info = artifact_row.iloc[0]
            stage_name = artifact_info.get('stage_name', '')
            artifact_name = artifact_info.get('artifact_name', '')
            
            # Map stage name to stage_id
            stage_id_mapping = {
                '0_drop_zone': 's0',
                '1_bronze': 's1', 
                '2_silver': 's2',
                '3_gold': 's3',
                '4_mart': 's4',
                '5_PBI_Model': 's5',
                '6_PBI_Reports': 's6'
            }
            stage_id = stage_id_mapping.get(stage_name, 's0')
            
            # Create new rows for this artifact using proper column structure
            new_rows = []
            for col_info in columns_info:
                # Use the established protected column structure from session context
                # IMPORTANT: Match exact column order from workbench setup
                new_row = {
                    'stage_id': stage_id,
                    'stage_name': stage_name,
                    'artifact_id': artifact_id,
                    'artifact_name': artifact_name,
                    'column_id': f"c{col_info['order']}",
                    'column_name': col_info['column_name'],
                    'data_type': col_info['data_type'],  # Correct position
                    'order': col_info['order'],           # Correct position
                    'column_business_name': '',  # Will be filled by AI
                    'column_group': '',  # Will be filled by cascade (lowercase as per session context)
                    'column_comment': ''  # Will be filled by AI
                }
                new_rows.append(new_row)
            
            # Remove existing columns for this artifact from Excel sheet directly
            self.remove_artifact_columns_from_sheet(artifact_id)
            
            # Append new rows using structure-preserving method
            new_df = pd.DataFrame(new_rows)
            
            # Use append method that preserves headers and formatting
            write_success = self.excel_utils.append_data_preserve_structure(self.workbook_path, "columns", new_df)
            if not write_success:
                self.logger.error(f"Failed to write columns for artifact {artifact_id} to Excel workbook")
                return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error adding columns to workbook: {str(e)}")
            return False
    
    def remove_artifact_columns_from_sheet(self, artifact_id: str) -> bool:
        """Remove existing columns for a specific artifact from the Excel sheet."""
        try:
            wb = load_workbook(self.workbook_path)
            if "columns" not in wb.sheetnames:
                return True  # No columns sheet, nothing to remove
            
            ws = wb["columns"]
            
            # Find artifact_id column index (should be column 3 based on structure)
            artifact_id_col = 3
            
            # Find and remove rows with matching artifact_id (start from bottom to avoid index issues)
            rows_to_delete = []
            for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                cell_value = ws.cell(row, artifact_id_col).value
                if cell_value == artifact_id:
                    rows_to_delete.append(row)
            
            # Delete rows from bottom to top to avoid index shifting
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)
            
            # Save workbook
            wb.save(self.workbook_path)
            return True
            
        except Exception as e:
            self.logger.error(f"Error removing artifact columns from sheet: {str(e)}")
            return False


# ANCHOR: Convenience Functions
def import_raw_files(workbook_path: str, project_path: str, source_folder: str = None) -> bool:
    """
    Convenience function to import raw CSV files into workbook.
    
    Args:
        workbook_path: Path to the Excel workbook
        project_path: Path to the project directory  
        source_folder: Path to sources folder (optional)
        
    Returns:
        bool: True if import successful
    """
    importer = RawFileImporter(workbook_path, project_path)
    return importer.import_assign_columns(source_folder)


if __name__ == "__main__":
    # Example usage
    print("Raw File Importer module loaded successfully")
    print("Use RawFileImporter class or import_raw_files() function")