"""
Excel Utilities
===============

Excel-specific operations and helpers.
"""

# ANCHOR: Imports and Dependencies

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# ANCHOR: ExcelUtils Class Definition

class ExcelUtils:
    """Excel workbook and worksheet utility functions."""
    
    # ANCHOR: Workbook Creation Methods
    @staticmethod
    def create_workbook_with_sheets(file_path: str, sheets_config: dict) -> bool:
        """
        Create Excel workbook with specified sheets and headers.
        
        Args:
            file_path: Path where to create the workbook
            sheets_config: Dictionary with sheet configurations
            
        Returns:
            bool: True if workbook created successfully
        """
        try:
            # Create new workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create each sheet
            for sheet_name, config in sheets_config.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                headers = config.get('headers', [])
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    # Format header cells
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add default data if provided
                default_data = config.get('default_data', [])
                for row_idx, row_data in enumerate(default_data, 2):
                    if isinstance(row_data, dict):
                        # Map dictionary data to columns
                        for col_idx, header in enumerate(headers, 1):
                            # Convert header to snake_case key for lookup
                            key = header.lower().replace(' ', '_')
                            value = row_data.get(key, '')
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    elif isinstance(row_data, list):
                        # Direct list data
                        for col_idx, value in enumerate(row_data, 1):
                            if col_idx <= len(headers):
                                ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
                
                # Create table formatting if there's data
                if len(headers) > 0:
                    last_row = max(2, len(default_data) + 1)
                    last_col = len(headers)
                    table_range = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"
                    
                    table = Table(displayName=f"Table_{sheet_name}", ref=table_range)
                    style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)
            
            # Save workbook
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error creating Excel workbook: {str(e)}")
            return False
    
    @staticmethod
    def read_sheet_data(file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        Read data from Excel sheet.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read
            
        Returns:
            pd.DataFrame: Sheet data as DataFrame
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Excel file not found: {file_path}")
                
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
            
        except Exception as e:
            print(f"Error reading Excel sheet: {str(e)}")
            return pd.DataFrame()
    
    @staticmethod
    def write_sheet_data(file_path: str, sheet_name: str, data: pd.DataFrame) -> bool:
        """
        Write data to Excel sheet.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to write
            data: DataFrame to write
            
        Returns:
            bool: True if write successful
        """
        try:
            # If file exists and is locked, try to close it via COM, update, then reopen
            if os.path.exists(file_path) and ExcelUtils._is_file_locked(file_path):
                print(f"Excel file appears to be open: {file_path}")
                try:
                    import win32com.client as win32
                    excel = win32.Dispatch("Excel.Application")
                    wb = None
                    for book in excel.Workbooks:
                        try:
                            if os.path.abspath(book.FullName).lower() == os.path.abspath(file_path).lower():
                                wb = book
                                break
                        except Exception:
                            continue
                    if wb:
                        print("Closing workbook in Excel via COM...")
                        wb.Close(SaveChanges=True)
                        # Wait a moment for file to unlock
                        import time
                        time.sleep(1)
                        # Confirm file is unlocked
                        if ExcelUtils._is_file_locked(file_path):
                            print("File still locked after COM close. Please close Excel manually.")
                            return False
                        print("Workbook closed. Proceeding with update...")
                        # Now update as normal
                        result = ExcelUtils.write_sheet_data(file_path, sheet_name, data)
                        # Reopen workbook in Excel for user
                        print("Reopening workbook in Excel...")
                        excel.Workbooks.Open(os.path.abspath(file_path))
                        return result
                    else:
                        print("Workbook not found in running Excel. Please close Excel manually.")
                        return False
                except ImportError:
                    print("pywin32 not installed. Please close Excel manually and try again.")
                    return False
                except Exception as e:
                    print(f"Error using COM to close Excel: {e}")
                    return False

            # Normal path when file not locked or doesn't exist
            if os.path.exists(file_path):
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Create new workbook
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)

            return True

        except PermissionError:
            print(f"Error: Permission denied writing to Excel file.")
            print(f"The file may be open in another application: {file_path}")
            print(f"Please close the file and try again.")
            return False
        except Exception as e:
            print(f"Error writing Excel sheet: {str(e)}")
            return False
    
    @staticmethod
    def validate_sheet_structure(file_path: str, sheet_name: str, expected_headers: list) -> bool:
        """
        Validate sheet has expected structure.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to validate
            expected_headers: List of expected header names
            
        Returns:
            bool: True if structure is valid
        """
        try:
            df = ExcelUtils.read_sheet_data(file_path, sheet_name)
            
            if df.empty:
                return False
                
            # Check if all expected headers are present
            actual_headers = df.columns.tolist()
            
            for expected_header in expected_headers:
                if expected_header not in actual_headers:
                    return False
                    
            return True
            
        except Exception as e:
            print(f"Error validating Excel sheet: {str(e)}")
            return False
    
    @staticmethod
    def get_sheet_names(file_path: str) -> list:
        """
        Get list of sheet names in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            list: List of sheet names
        """
        try:
            if not os.path.exists(file_path):
                return []
                
            xl_file = pd.ExcelFile(file_path)
            return xl_file.sheet_names
            
        except Exception as e:
            print(f"Error getting sheet names: {str(e)}")
            return []

    @staticmethod
    def _is_file_locked(file_path: str) -> bool:
        """
        Check if file is locked (open in another application).
        
        Args:
            file_path: Path to file to check
            
        Returns:
            bool: True if file is locked
        """
        try:
            # Try to open file in write mode
            with open(file_path, 'r+b'):
                pass
            return False
        except (IOError, OSError, PermissionError):
            return True

    @staticmethod
    def _write_via_com(file_path: str, sheet_name: str, data: pd.DataFrame) -> bool:
        """
        Write data to an open Excel workbook using COM (pywin32). This allows updating the workbook
        even when Excel has the file open.
        """
        try:
            # Import locally to avoid hard dependency at module import time
            import win32com.client as win32

            excel = win32.Dispatch("Excel.Application")

            # Try to find an already opened workbook
            wb = None
            for book in excel.Workbooks:
                try:
                    if os.path.abspath(book.FullName).lower() == os.path.abspath(file_path).lower():
                        wb = book
                        break
                except Exception:
                    continue

            if wb is None:
                # Open without ReadOnly to allow saving; ignore alerts
                wb = excel.Workbooks.Open(os.path.abspath(file_path), ReadOnly=False)

            # Try to get the worksheet; if it doesn't exist, add it
            try:
                ws = wb.Worksheets(sheet_name)
            except Exception:
                ws = wb.Worksheets.Add()
                ws.Name = sheet_name

            # Prepare data as a 2D list (headers + rows)
            headers = list(data.columns)
            rows = data.fillna("").values.tolist()
            table = [headers] + rows

            # Determine target range size
            n_rows = len(table)
            n_cols = len(headers)

            # Write values starting at A1
            start_cell = ws.Range("A1")
            end_cell = ws.Cells(n_rows, n_cols)
            write_range = ws.Range(start_cell, end_cell)

            # COM accepts a tuple of tuples
            write_range.Value = tuple(tuple(row) for row in table)

            # Save workbook via COM
            wb.Save()
            return True
        except ModuleNotFoundError:
            raise
        except Exception as e:
            # Re-raise to allow fallback to xlwings
            raise

    @staticmethod
    def _write_via_xlwings(file_path: str, sheet_name: str, data: pd.DataFrame) -> bool:
        """
        Write data to an open Excel workbook using xlwings.
        """
        try:
            import xlwings as xw

            # Try to attach to active app if present
            app = None
            if xw.apps:
                app = xw.apps.active
            if app is None:
                app = xw.App(visible=False)

            # Try to connect to the workbook if open, otherwise open it
            try:
                book = xw.Book(os.path.abspath(file_path))
            except Exception:
                book = app.books.open(os.path.abspath(file_path))

            # Ensure sheet exists
            try:
                sheet = book.sheets[sheet_name]
            except Exception:
                sheet = book.sheets.add(sheet_name)

            # Write dataframe (xlwings handles headers)
            sheet.range("A1").options(index=False).value = data

            # Save workbook
            book.save()
            return True
        except ModuleNotFoundError:
            raise
        except Exception as e:
            raise
