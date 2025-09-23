"""
Workbench Configuration Setup
============================

Sets up default configuration files for workbench functionality.
Creates workbench_configuration_*.xlsx files with complete structure from AdventureWorks.
"""

# ANCHOR: Imports and Dependencies
import pandas as pd
from pathlib import Path
import sys

# Import utilities
current_dir = Path(__file__).parent
src_dir = current_dir.parent
sys.path.insert(0, str(src_dir))

from utils.c_workbench_excel_utils import ExcelUtils
from utils.z_logger import Logger

# ANCHOR: WorkbenchConfigurationManager Class
class WorkbenchConfigurationManager:
    """Manages workbench configuration setup and defaults."""
    
    def __init__(self):
        self.excel_utils = ExcelUtils()
        self.logger = Logger()
    
    # ANCHOR: Project Configuration Methods
    def create_project_config_file(self, config_path: str, project_name: str) -> bool:
        """
        Create project-specific workbench configuration file.
        Based on complete AdventureWorks structure.
        
        Args:
            config_path: Path where to create the configuration file
            project_name: Name of the project (for naming only, structure stays same)
            
        Returns:
            bool: True if configuration file created successfully
        """
        try:
            # Create all sheets with complete AdventureWorks structure
            data_type_mappings = self._create_data_type_mappings()
            technical_columns = self._create_technical_columns()
            stage_configuration = self._create_stage_configuration()
            relation_types = self._create_relation_types()
            
            # Write to Excel file with multiple sheets in the correct order
            with pd.ExcelWriter(config_path, engine='openpyxl') as writer:
                # Write sheets in logical order: stages → technical_columns → relations → data_mappings
                stage_configuration.to_excel(writer, sheet_name='stages', index=False)
                technical_columns.to_excel(writer, sheet_name='technical_columns', index=False)
                relation_types.to_excel(writer, sheet_name='relations', index=False)
                data_type_mappings.to_excel(writer, sheet_name='data_mappings', index=False)
                
                # Apply formatting with light grey for lookup columns
                self._apply_sheet_formatting(writer)
            
            self.logger.info(f"Created project-specific workbench configuration file: {config_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to create configuration file: {str(e)}")
            return False
    
    def create_default_config_file(self, config_path: str) -> bool:
        """Create default workbench configuration file."""
        return self.create_project_config_file(config_path, "DefaultProject")
    
    def _create_data_type_mappings(self) -> pd.DataFrame:
        """Create data type mappings from AdventureWorks structure."""
        # Complete AdventureWorks DataTypeMappings data
        datatypemappings_data = [
            {'source': 'INT', 'sql_server': 'INT', 'databricks': 'INT', 'power_bi': 'INT64', 'notes': 'Standard integer'},
            {'source': 'BIGINT', 'sql_server': 'BIGINT', 'databricks': 'BIGINT', 'power_bi': 'INT64', 'notes': 'Large integer'},
            {'source': 'SMALLINT', 'sql_server': 'SMALLINT', 'databricks': 'SMALLINT', 'power_bi': 'INT64', 'notes': 'Small integer'},
            {'source': 'TINYINT', 'sql_server': 'TINYINT', 'databricks': 'TINYINT', 'power_bi': 'INT64', 'notes': 'Tiny integer'},
            {'source': 'BIT', 'sql_server': 'BIT', 'databricks': 'BOOLEAN', 'power_bi': 'Boolean', 'notes': 'Boolean flag'},
            {'source': 'DECIMAL', 'sql_server': 'DECIMAL(18,2)', 'databricks': 'DECIMAL(18,2)', 'power_bi': 'Decimal', 'notes': 'Precise decimal'},
            {'source': 'NUMERIC', 'sql_server': 'NUMERIC(18,2)', 'databricks': 'NUMERIC(18,2)', 'power_bi': 'Decimal', 'notes': 'Numeric with precision'},
            {'source': 'FLOAT', 'sql_server': 'FLOAT', 'databricks': 'FLOAT', 'power_bi': 'Double', 'notes': 'Floating point'},
            {'source': 'REAL', 'sql_server': 'REAL', 'databricks': 'REAL', 'power_bi': 'Double', 'notes': 'Real number'},
            {'source': 'MONEY', 'sql_server': 'MONEY', 'databricks': 'DECIMAL(19,4)', 'power_bi': 'Decimal', 'notes': 'Currency values'},
            {'source': 'CHAR', 'sql_server': 'CHAR(255)', 'databricks': 'CHAR(255)', 'power_bi': 'String', 'notes': 'Fixed length char'},
            {'source': 'VARCHAR', 'sql_server': 'VARCHAR(MAX)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'Variable length string'},
            {'source': 'TEXT', 'sql_server': 'NVARCHAR(MAX)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'Large text'},
            {'source': 'NCHAR', 'sql_server': 'NCHAR(255)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'Unicode fixed char'},
            {'source': 'NVARCHAR', 'sql_server': 'NVARCHAR(MAX)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'Unicode variable string'},
            {'source': 'NTEXT', 'sql_server': 'NVARCHAR(MAX)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'Unicode large text'},
            {'source': 'DATE', 'sql_server': 'DATE', 'databricks': 'DATE', 'power_bi': 'Date', 'notes': 'Date only'},
            {'source': 'DATETIME', 'sql_server': 'DATETIME', 'databricks': 'TIMESTAMP', 'power_bi': 'DateTime', 'notes': 'Date and time'},
            {'source': 'DATETIME2', 'sql_server': 'DATETIME2', 'databricks': 'TIMESTAMP', 'power_bi': 'DateTime', 'notes': 'Enhanced datetime'},
            {'source': 'SMALLDATETIME', 'sql_server': 'SMALLDATETIME', 'databricks': 'TIMESTAMP', 'power_bi': 'DateTime', 'notes': 'Small datetime'},
            {'source': 'TIME', 'sql_server': 'TIME', 'databricks': 'TIME', 'power_bi': 'Time', 'notes': 'Time only'},
            {'source': 'TIMESTAMP', 'sql_server': 'DATETIME2', 'databricks': 'TIMESTAMP', 'power_bi': 'DateTime', 'notes': 'Timestamp'},
            {'source': 'BINARY', 'sql_server': 'BINARY(8000)', 'databricks': 'BINARY', 'power_bi': 'Binary', 'notes': 'Fixed binary'},
            {'source': 'VARBINARY', 'sql_server': 'VARBINARY(MAX)', 'databricks': 'BINARY', 'power_bi': 'Binary', 'notes': 'Variable binary'},
            {'source': 'IMAGE', 'sql_server': 'VARBINARY(MAX)', 'databricks': 'BINARY', 'power_bi': 'Binary', 'notes': 'Image/blob data'},
            {'source': 'UNIQUEIDENTIFIER', 'sql_server': 'UNIQUEIDENTIFIER', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'GUID/UUID'},
            {'source': 'XML', 'sql_server': 'XML', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'XML data'},
            {'source': 'JSON', 'sql_server': 'NVARCHAR(MAX)', 'databricks': 'STRING', 'power_bi': 'String', 'notes': 'JSON data'}
        ]
        
        return pd.DataFrame(datatypemappings_data)
        
    def _create_technical_columns(self) -> pd.DataFrame:
        """Create technical columns from AdventureWorks structure with stage_name lookup.
        Structure matches config_test.xlsx - DO NOT CHANGE WITHOUT PERMISSION."""
        # Complete AdventureWorks TechnicalColumns data with actual stage names for lookup
        technicalcolumns_data = [
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__SourceSystem', 'data_type': 'STRING', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Source system identifier'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__SourceFileName', 'data_type': 'STRING', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Original source file name'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__SourceFilePath', 'data_type': 'STRING', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Source file path'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__bronze_insert_dt', 'data_type': 'TIMESTAMP', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Bronze insertion timestamp'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__bronzePartition_InsertYear', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Partition year for bronze data'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__bronzePartition_InsertMonth', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Partition month for bronze data'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'column_name': '__bronzePartition_insertDate', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Partition date for bronze data'},
            {'stage_id': 's2', 'stage_name': '2_silver', 'column_name': '__silver_last_changed_dt', 'data_type': 'TIMESTAMP', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Silver last changed timestamp'},
            {'stage_id': 's2', 'stage_name': '2_silver', 'column_name': '__silverPartition_xxxYear', 'data_type': 'INT', 'include_in_tech_fields': False, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Silver partition year (xxx = business date field)'},
            {'stage_id': 's2', 'stage_name': '2_silver', 'column_name': '__silverPartition_xxxMonth', 'data_type': 'INT', 'include_in_tech_fields': False, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Silver partition month (xxx = business date field)'},
            {'stage_id': 's2', 'stage_name': '2_silver', 'column_name': '__silverPartition_xxxDate', 'data_type': 'INT', 'include_in_tech_fields': False, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Silver partition date (xxx = business date field)'},
            {'stage_id': 's3', 'stage_name': '3_gold', 'column_name': '__gold_last_changed_dt', 'data_type': 'TIMESTAMP', 'include_in_tech_fields': True, 'take_to_next_level': 1, 'artifact_type_specific': 'all', 'description': 'Gold last changed timestamp'},
            {'stage_id': 's3', 'stage_name': '3_gold', 'column_name': '__goldPartition_XXXYear', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Gold partition year (XXX = business date field)'},
            {'stage_id': 's3', 'stage_name': '3_gold', 'column_name': '__goldPartition_XXXMonth', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Gold partition month (XXX = business date field)'},
            {'stage_id': 's3', 'stage_name': '3_gold', 'column_name': '__goldPartition_XXXDate', 'data_type': 'INT', 'include_in_tech_fields': True, 'take_to_next_level': 0, 'artifact_type_specific': 'all', 'description': 'Gold partition date (XXX = business date field)'}
        ]
        
        return pd.DataFrame(technicalcolumns_data)
        
    def _create_stage_configuration(self) -> pd.DataFrame:
        """Create stage configuration from AdventureWorks structure."""
        # Complete AdventureWorks StageConfiguration data
        stageconfiguration_data = [
            {'stage_id': 's0', 'stage_name': '0_drop_zone', 'platform': 'databricks', 'artifact_side': 'source', 'description': 'Drop zone for raw data files', 'processing_notes': 'Raw data ingestion, minimal processing'},
            {'stage_id': 's1', 'stage_name': '1_bronze', 'platform': 'databricks', 'artifact_side': 'source', 'description': 'Bronze layer for raw structured data', 'processing_notes': 'Raw storage with basic structure'},
            {'stage_id': 's2', 'stage_name': '2_silver', 'platform': 'databricks', 'artifact_side': 'source', 'description': 'Silver layer for cleaned and validated data', 'processing_notes': 'Cleaned and dedublicated source data'},
            {'stage_id': 's3', 'stage_name': '3_gold', 'platform': 'databricks', 'artifact_side': 'business', 'description': 'Gold layer for business-ready data', 'processing_notes': 'Business ready, conformed dimensions'},
            {'stage_id': 's4', 'stage_name': '4_mart', 'platform': 'databricks', 'artifact_side': 'business', 'description': 'Data marts for specific business domains', 'processing_notes': 'Analytical marts ,products , domains'},
            {'stage_id': 's5', 'stage_name': '5_PBI_Model', 'platform': 'power bi', 'artifact_side': 'business', 'description': 'Power BI semantic model layer', 'processing_notes': 'Power BI optimized model'},
            {'stage_id': 's6', 'stage_name': '6_PBI_Reports', 'platform': 'power bi', 'artifact_side': 'business', 'description': 'Power BI reports and dashboards', 'processing_notes': 'Report layer definitions'}
        ]
        
        return pd.DataFrame(stageconfiguration_data)
        
    def _create_relation_types(self) -> pd.DataFrame:
        """Create relation types from AdventureWorks structure.
        Structure matches config_test.xlsx - DO NOT CHANGE WITHOUT PERMISSION."""
        # Complete AdventureWorks RelationTypes data
        relationtypes_data = [
            {'relation_type': 'main', 'description': 'Full column propagation with technical fields', 'processing_logic': 'Context-aware processing based on artifact type and stage transition', 'field_limit': 'No limit', 'use_cases': 'Standard column cascading between stages'},
            {'relation_type': 'get_key', 'description': 'Dimension key propagation for fact tables', 'processing_logic': 'Extracts surrogate keys (SKs) and business keys (BKs) only', 'field_limit': 'Keys only', 'use_cases': 'Foreign key relationships in fact tables'},
            {'relation_type': 'lookup', 'description': 'Limited column lookup with priority-based selection', 'processing_logic': 'Priority order: SKs → BKs → Attributes, configurable limit', 'field_limit': '3 (default)', 'use_cases': 'Reference lookups and denormalization'},
            {'relation_type': 'pbi', 'description': 'Power BI specific minimal cascading', 'processing_logic': 'Keys and facts only for analytical performance', 'field_limit': 'Keys + Facts', 'use_cases': 'Power BI model optimization'}
        ]
        
        return pd.DataFrame(relationtypes_data)
        
    def _apply_sheet_formatting(self, writer):
        """Apply formatting to the Excel sheets, including light grey for lookup columns."""
        try:
            from openpyxl.styles import PatternFill
            
            # Light grey fill for lookup columns
            light_grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Get the workbook
            workbook = writer.book
            
            # Format technical_columns sheet - stage_name column (column B) is a lookup
            if 'technical_columns' in workbook.sheetnames:
                ws = workbook['technical_columns']
                # Apply light grey to stage_name column (column B)
                for row in range(1, ws.max_row + 1):
                    ws[f'B{row}'].fill = light_grey_fill
                    
        except ImportError:
            # openpyxl styling not available, skip formatting
            self.logger.warning("openpyxl styling not available, skipping column formatting")
        except Exception as e:
            self.logger.warning(f"Could not apply sheet formatting: {str(e)}")

def main():
    """Create default configuration file."""
    config_manager = WorkbenchConfigurationManager()
    
    # Create in current directory for testing
    config_path = Path(__file__).parent / "workbench_configuration_default.xlsx"
    
    success = config_manager.create_default_config_file(str(config_path))
    
    if success:
        print(f"Successfully created configuration file: {config_path}")
    else:
        print("Failed to create configuration file")

if __name__ == "__main__":
    main()
