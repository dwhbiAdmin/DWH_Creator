"""
Raw Files Enhancement Module
===========================

Module for enhancing imported raw files with:
1. AI-generated column comments
2. AI-generated artifact comments  
3. Deterministic primary key detection
4. AI-generated business names for columns

This module consolidates all post-import enhancement operations.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple, Any
import re
from collections import Counter

# Import project utilities
from utils.logger import Logger
from utils.B_worksheet_config_manager import ConfigManager
from utils.B_excel_utils import ExcelUtils
from backend.Y_ai_manager import AIWorkbenchManager
from utils.Y_ai_comment_generator import AICommentGenerator


class RawFilesEnhancer:
    """
    Main class for enhancing imported raw files.
    
    Provides comprehensive enhancement including:
    - AI-generated column comments
    - AI-generated artifact comments
    - Deterministic primary key detection
    - AI-generated business names
    """
    
    def __init__(self, workbook_path: str, api_key: Optional[str] = None):
        """
        Initialize the Raw Files Enhancer.
        
        Args:
            workbook_path: Path to the workbook file
            api_key: Optional API key for AI services
        """
        self.logger = Logger()
        self.config_manager = ConfigManager()
        self.excel_utils = ExcelUtils()
        self.workbook_path = workbook_path
        self.api_key = api_key  # Store API key for AI generators
        
        # Initialize AI manager if API key provided
        self.ai_manager = None
        if api_key:
            try:
                self.ai_manager = AIWorkbenchManager(workbook_path, api_key)
                self.logger.info("AI manager initialized successfully")
            except Exception as e:
                self.logger.warning(f"Could not initialize AI manager: {str(e)}")
        
        # Configuration for primary key detection
        self.pk_detection_config = {
            "min_uniqueness_ratio": 0.95,  # Minimum uniqueness for PK consideration
            "max_null_ratio": 0.05,        # Maximum null percentage for PK
        }
        
        # Patterns for identifying key columns by name
        self.key_patterns = [
            r".*_?id$", r".*_?key$", r".*_?pk$", 
            r"^id$", r"^key$", r"^pk$",
            r".*_?identifier$", r".*_?number$"
        ]
        
        self.logger.info("RawFilesEnhancer initialized successfully")
    
    def enhance_all_raw_files(self) -> bool:
        """
        Perform all enhancement operations on imported raw files.
        
        Returns:
            bool: True if all enhancements completed successfully
        """
        self.logger.info("Starting comprehensive enhancement of raw files")
        
        try:
            # Step 1: Generate AI comments for artifacts
            self.logger.info("Step 1: Generating AI comments for artifacts...")
            artifact_success = self.generate_artifact_comments()
            
            # Step 2: Generate AI comments for columns
            self.logger.info("Step 2: Generating AI comments for columns...")
            column_comments_success = self.generate_column_comments()
            
            # Step 3: Determine primary keys deterministically
            self.logger.info("Step 3: Determining primary keys...")
            pk_success = self.determine_primary_keys()
            
            # Step 4: Generate AI business names for columns
            self.logger.info("Step 4: Generating business names for columns...")
            business_names_success = self.generate_business_names()
            
            # Check overall success
            overall_success = all([
                artifact_success, 
                column_comments_success, 
                pk_success, 
                business_names_success
            ])
            
            if overall_success:
                self.logger.info("✅ All enhancement operations completed successfully!")
            else:
                self.logger.warning("⚠️ Some enhancement operations had issues. Check logs for details.")
            
            return overall_success
            
        except Exception as e:
            self.logger.error(f"Error during enhancement process: {str(e)}")
            return False
    
    def generate_artifact_comments(self) -> bool:
        """Generate AI comments for artifacts."""
        try:
            if not self.ai_manager:
                self.logger.warning("AI manager not available. Skipping artifact comments.")
                return True  # Not a failure, just skip
            
            success = self.ai_manager.generate_artifact_comments()
            if success:
                self.logger.info("✅ Artifact comments generated successfully")
            else:
                self.logger.warning("⚠️ Some artifact comments could not be generated")
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error generating artifact comments: {str(e)}")
            return False
    
    def generate_column_comments(self) -> bool:
        """Generate AI comments for columns."""
        try:
            if not self.ai_manager:
                self.logger.warning("AI manager not available. Skipping column comments.")
                return True  # Not a failure, just skip
            
            success = self.ai_manager.generate_column_comments()
            if success:
                self.logger.info("✅ Column comments generated successfully")
            else:
                self.logger.warning("⚠️ Some column comments could not be generated")
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error generating column comments: {str(e)}")
            return False
    
    def determine_primary_keys(self) -> bool:
        """Deterministically identify and set primary keys for all column sheets."""
        try:
            # Get all column sheet names
            column_sheets = self._get_column_sheet_names()
            
            if not column_sheets:
                self.logger.warning("No column sheets found to analyze for primary keys")
                return True
            
            updated_sheets = 0
            
            for sheet_name in column_sheets:
                try:
                    # Read the column sheet
                    columns_df = self.excel_utils.read_sheet_data(self.workbook_path, sheet_name)
                    
                    if columns_df.empty:
                        self.logger.warning(f"Sheet {sheet_name} is empty, skipping")
                        continue
                    
                    # Analyze and update primary keys for this sheet
                    pk_candidates = self._analyze_primary_key_candidates(columns_df, sheet_name)
                    
                    if pk_candidates:
                        # Update the column_group for primary key columns
                        updated_df = self._update_primary_key_groups(columns_df, pk_candidates)
                        
                        # Write back to Excel - PRESERVE FORMATTING (frozen headers, filters)
                        write_success = self.excel_utils.write_sheet_data_preserve_formatting(
                            self.workbook_path, sheet_name, updated_df
                        )
                        
                        if write_success:
                            updated_sheets += 1
                            pk_cols = ", ".join(pk_candidates)
                            self.logger.info(f"✅ Primary key(s) identified for {sheet_name}: {pk_cols}")
                        else:
                            self.logger.error(f"Failed to write primary key updates to {sheet_name}")
                    else:
                        self.logger.info(f"No clear primary key candidates found for {sheet_name}")
                
                except Exception as e:
                    self.logger.error(f"Error analyzing primary keys for {sheet_name}: {str(e)}")
                    continue
            
            self.logger.info(f"✅ Primary key analysis completed. Updated {updated_sheets} sheets.")
            return True
            
        except Exception as e:
            self.logger.error(f"Error in primary key determination: {str(e)}")
            return False
    
    def generate_business_names(self) -> bool:
        """Generate AI business names for columns using dedicated AI module."""
        try:
            # Try to initialize AI comment generator if we have an API key
            ai_generator = None
            if hasattr(self, 'api_key') and self.api_key:
                try:
                    ai_generator = AICommentGenerator(self.api_key)
                    if not ai_generator.is_available():
                        ai_generator = None
                except Exception as e:
                    self.logger.warning(f"Could not initialize AI comment generator: {str(e)}")
                    ai_generator = None
            
            if not ai_generator:
                self.logger.warning("AI generator not available. Generating simple business names.")
                return self._generate_simple_business_names()
            
            # Get column sheets to process
            column_sheets = self._get_column_sheet_names()
            if not column_sheets:
                self.logger.warning("No column sheets found for business name generation")
                return False
            
            updated_sheets = 0
            for sheet_name in column_sheets:
                try:
                    # Read current column data
                    columns_df = self.excel_utils.read_sheet_data(self.workbook_path, sheet_name)
                    
                    if columns_df.empty:
                        continue
                    
                    # Prepare data for AI processing
                    columns_data = []
                    for _, row in columns_df.iterrows():
                        column_name = row.get('column_name', row.get('Column Name', ''))
                        data_type = row.get('data_type', row.get('Data Type', ''))
                        current_business_name = row.get('column_business_name', row.get('Column Business Name', ''))
                        
                        # Only process if business name is empty and we have column name
                        if column_name and (pd.isna(current_business_name) or current_business_name.strip() == ''):
                            columns_data.append({
                                'column_name': column_name,
                                'data_type': data_type
                            })
                    
                    if not columns_data:
                        continue  # No columns need business names
                    
                    # Generate AI business names in batch
                    self.logger.info(f"Generating AI business names for {len(columns_data)} columns in {sheet_name}")
                    ai_business_names = ai_generator.generate_business_names_batch(columns_data)
                    
                    # Update the dataframe
                    updated_df = columns_df.copy()
                    for idx, row in updated_df.iterrows():
                        column_name = row.get('column_name', row.get('Column Name', ''))
                        
                        if column_name in ai_business_names:
                            # Update the correct column name format
                            if 'column_business_name' in updated_df.columns:
                                updated_df.at[idx, 'column_business_name'] = ai_business_names[column_name]
                            elif 'Column Business Name' in updated_df.columns:
                                updated_df.at[idx, 'Column Business Name'] = ai_business_names[column_name]
                    
                    # Write back to Excel - PRESERVE FORMATTING (frozen headers, filters)
                    write_success = self.excel_utils.write_sheet_data_preserve_formatting(
                        self.workbook_path, sheet_name, updated_df
                    )
                    
                    if write_success:
                        updated_sheets += 1
                        self.logger.info(f"✅ AI business names updated for {sheet_name}")
                
                except Exception as e:
                    self.logger.error(f"Error generating AI business names for {sheet_name}: {str(e)}")
                    continue
            
            if updated_sheets > 0:
                self.logger.info(f"✅ AI business names generated for {updated_sheets} sheets")
                return True
            else:
                self.logger.warning("⚠️ No AI business names were generated, falling back to simple names")
                return self._generate_simple_business_names()
            
        except Exception as e:
            self.logger.error(f"Error generating AI business names: {str(e)}")
            # Fallback to simple business names
            return self._generate_simple_business_names()
    
    def _get_column_sheet_names(self) -> List[str]:
        """Get all column sheet names from the workbook."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.workbook_path, read_only=True)
            
            # Look for both individual column sheets (Columns1, Columns2, etc.) and consolidated sheet (columns)
            column_sheets = []
            
            # Check for individual column sheets (from CSV imports)
            individual_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith('Columns') and sheet != 'Columns']
            column_sheets.extend(individual_sheets)
            
            # Check for consolidated columns sheet (standard workbook)
            if 'columns' in wb.sheetnames:
                column_sheets.append('columns')
            
            wb.close()
            
            return column_sheets
            
        except Exception as e:
            self.logger.error(f"Error getting column sheet names: {str(e)}")
            return []
    
    def _analyze_primary_key_candidates(self, columns_df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Analyze a column sheet to identify primary key candidates."""
        try:
            # First, get the actual data from the source to analyze uniqueness
            # For now, we'll use heuristics based on column names and types
            
            pk_candidates = []
            
            # Check each column for primary key characteristics
            for _, row in columns_df.iterrows():
                # Try both column name formats
                column_name = row.get('column_name', row.get('Column Name', ''))
                data_type = row.get('data_type', row.get('Data Type', ''))
                column_group = row.get('column_group', row.get('Column Group', ''))
                
                # Skip if already marked as primary key
                if column_group == 'primary_key':
                    continue
                
                # Check name patterns
                col_lower = column_name.lower()
                is_key_pattern = any(re.match(pattern, col_lower) for pattern in self.key_patterns)
                
                # Check data type (prefer integer or string IDs)
                is_appropriate_type = any(dt in data_type.upper() for dt in ['INT', 'BIGINT', 'VARCHAR', 'STRING'])
                
                # Score the candidate
                score = 0
                if is_key_pattern:
                    score += 3
                if is_appropriate_type:
                    score += 2
                if 'id' in col_lower:
                    score += 2
                if col_lower in ['id', 'key', 'pk']:
                    score += 3
                
                # Consider as primary key candidate if score is high enough
                if score >= 4:
                    pk_candidates.append(column_name)
                    self.logger.info(f"Primary key candidate found: {column_name} (score: {score})")
            
            # If multiple candidates, prefer the one with the highest score/best name
            if len(pk_candidates) > 1:
                # Prefer columns with simpler names
                preferred = None
                for candidate in pk_candidates:
                    col_lower = candidate.lower()
                    if col_lower in ['id', 'key', 'pk']:
                        preferred = candidate
                        break
                    elif col_lower.endswith('_id') and not preferred:
                        preferred = candidate
                
                if preferred:
                    pk_candidates = [preferred]
                else:
                    # Take the first one
                    pk_candidates = pk_candidates[:1]
            
            return pk_candidates
            
        except Exception as e:
            self.logger.error(f"Error analyzing primary key candidates for {sheet_name}: {str(e)}")
            return []
    
    def _update_primary_key_groups(self, columns_df: pd.DataFrame, pk_candidates: List[str]) -> pd.DataFrame:
        """Update the column_group for primary key columns."""
        updated_df = columns_df.copy()
        
        # Update the Column Group for primary key candidates
        mask = updated_df['Column Name'].isin(pk_candidates)
        updated_df.loc[mask, 'Column Group'] = 'primary_key'
        
        return updated_df
    
    def _generate_simple_business_names(self) -> bool:
        """Generate simple business names without AI (fallback method)."""
        try:
            # Get all column sheet names
            column_sheets = self._get_column_sheet_names()
            
            if not column_sheets:
                self.logger.warning("No column sheets found for business name generation")
                return True
            
            updated_sheets = 0
            
            for sheet_name in column_sheets:
                try:
                    # Read the column sheet
                    columns_df = self.excel_utils.read_sheet_data(self.workbook_path, sheet_name)
                    
                    if columns_df.empty:
                        continue
                    
                    # Generate simple business names
                    updated_df = columns_df.copy()
                    
                    for idx, row in updated_df.iterrows():
                        # Try both column name formats (with spaces/underscores)
                        column_name = row.get('column_name', row.get('Column Name', ''))
                        current_business_name = row.get('column_business_name', row.get('Column Business Name', ''))
                        
                        # Only generate if business name is empty
                        if pd.isna(current_business_name) or current_business_name.strip() == '':
                            business_name = self._create_simple_business_name(column_name)
                            # Update the correct column name format
                            if 'column_business_name' in updated_df.columns:
                                updated_df.at[idx, 'column_business_name'] = business_name
                            elif 'Column Business Name' in updated_df.columns:
                                updated_df.at[idx, 'Column Business Name'] = business_name
                    
                    # Write back to Excel - PRESERVE FORMATTING (frozen headers, filters)
                    write_success = self.excel_utils.write_sheet_data_preserve_formatting(
                        self.workbook_path, sheet_name, updated_df
                    )
                    
                    if write_success:
                        updated_sheets += 1
                
                except Exception as e:
                    self.logger.error(f"Error generating business names for {sheet_name}: {str(e)}")
                    continue
            
            self.logger.info(f"✅ Simple business names generated for {updated_sheets} sheets")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating simple business names: {str(e)}")
            return False
    
    def _create_simple_business_name(self, column_name: str) -> str:
        """Create a simple business name from column name in snake_case format."""
        # Start with the original column name in lowercase
        business_name = column_name.lower()
        
        # Enhanced abbreviations and common patterns for snake_case
        replacements = {
            # IDs and Keys
            'cust_id': 'customer_id',
            'cust_': 'customer_',
            'prod_id': 'product_id',
            'prod_': 'product_',
            'ord_id': 'order_id',
            'ord_': 'order_',
            'emp_id': 'employee_id',
            'emp_': 'employee_',
            'acct_id': 'account_id',
            'acct_': 'account_',
            'clr_id': 'color_id',
            'clr_': 'color_',
            
            # Names and descriptions
            'cust_name': 'customer_name',
            'prod_name': 'product_name',
            'emp_name': 'employee_name',
            
            # Dates and Times
            '_dt': '_date',
            '_ts': '_timestamp',
            '_tm': '_time',
            
            # Quantities and Amounts
            'qty': 'quantity',
            '_amt': '_amount',
            '_num': '_number',
            '_cnt': '_count',
            
            # Descriptions and Text
            '_desc': '_description',
            '_nm': '_name',
            '_addr': '_address',
            '_cd': '_code',
            '_sts': '_status',
            '_cat': '_category',
            
            # Common business terms
            'm_or_f': 'gender',
            '_ref': '_reference',
            '_val': '_value',
            '_pct': '_percentage'
        }
        
        # Apply replacements
        for old, new in replacements.items():
            business_name = business_name.replace(old, new)
        
        return business_name


def enhance_raw_files(workbook_path: str, api_key: Optional[str] = None) -> bool:
    """
    Convenience function to enhance raw files.
    
    Args:
        workbook_path: Path to the workbook file
        api_key: Optional API key for AI services
        
    Returns:
        bool: True if enhancement completed successfully
    """
    enhancer = RawFilesEnhancer(workbook_path, api_key)
    return enhancer.enhance_all_raw_files()