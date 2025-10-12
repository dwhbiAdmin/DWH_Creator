#!/usr/bin/env python3
"""
Workbench Setup Module for DWH Creator
Created: September 22, 2025
Purpose: Handle setup of workbench sheets (stages, artifacts, columns) with embedded default structure
Similar to workbench_configuration_setup.py but for actual workbench files

This module creates and manages the workbench Excel files that contain:
- Stages sheet: Stage definitions and metadata
- Artifacts sheet: Artifact definitions for each stage  
- Columns sheet: Column definitions and cascading rules

Structure is based on the established workbench patterns and will NOT be changed without permission.
"""

import os
import logging
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Any

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('DWH_Creator')


class WorkbenchSetupManager:
    """
    Manages creation and setup of workbench Excel files with embedded default structure.
    Handles the 7-sheet integrated workbook creation for project initialization.
    """
    
    def __init__(self):
        """Initialize the WorkbenchSetupManager."""
        self.default_structure_loaded = True
        logger.info("WorkbenchSetupManager initialized with embedded default structure")
    
    def create_project_workbench_file(self, file_path: str, project_name: str) -> bool:
        """
        Create a new workbench Excel file with default structure.
        
        Args:
            file_path (str): Path where the workbench file will be created
            project_name (str): Name of the project for metadata
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create the workbench sheets with embedded data
            stages_df = self._create_stages_sheet()
            artifacts_df = self._create_artifacts_sheet()
            columns_df = self._create_columns_sheet()
            
            # Write to Excel with proper sheet order
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write sheets in logical order: stages → artifacts → columns
                stages_df.to_excel(writer, sheet_name='stages', index=False)
                artifacts_df.to_excel(writer, sheet_name='artifacts', index=False)  
                columns_df.to_excel(writer, sheet_name='columns', index=False)
                
                # Apply formatting
                self._apply_workbench_formatting(writer)
            
            logger.info(f"Created project-specific workbench file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to create workbench file {file_path}: {str(e)}")
            return False
    
    def create_default_workbench_file(self, config_path: str = "workbench_default.xlsx") -> bool:
        """
        Create a default workbench file in the current directory.
        
        Args:
            config_path (str): Path for the default workbench file
            
        Returns:
            bool: True if successful, False otherwise
        """
        return self.create_project_workbench_file(config_path, "DefaultWorkbench")
    
    def create_integrated_workbench_file(self, file_path: str, project_name: str) -> bool:
        """
        Create a new integrated workbench Excel file with 7 sheets (3 visible + 4 hidden config).
        This is the new 7-sheet integrated approach with all configuration embedded.
        
        Args:
            file_path (str): Path where the workbench file will be created
            project_name (str): Name of the project for metadata
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create the main workbench sheets
            stages_df = self._create_stages_sheet()
            artifacts_df = self._create_artifacts_sheet()
            columns_df = self._create_columns_sheet()
            
            # Create the 4 configuration sheets with proper names
            conf_1_stages_df = self._create_config_metadata_sheet(project_name)
            conf_2_technical_columns_df = self._create_config_settings_sheet()
            conf_3_relations_df = self._create_config_validation_sheet()
            conf_4_data_mappings_df = self._create_config_templates_sheet()
            
            # Write to Excel with proper sheet order (visible first, then config)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write 3 visible sheets first
                stages_df.to_excel(writer, sheet_name='stages', index=False)
                artifacts_df.to_excel(writer, sheet_name='artifacts', index=False)  
                columns_df.to_excel(writer, sheet_name='columns', index=False)
                
                # Write 4 configuration sheets with conf_ prefixes (kept visible)
                conf_1_stages_df.to_excel(writer, sheet_name='conf_1_stages', index=False)
                conf_2_technical_columns_df.to_excel(writer, sheet_name='conf_2_technical_columns', index=False)
                conf_3_relations_df.to_excel(writer, sheet_name='conf_3_relations', index=False)
                conf_4_data_mappings_df.to_excel(writer, sheet_name='conf_4_data_mappings', index=False)
                
                # Apply formatting to all sheets
                self._apply_workbench_formatting(writer)
                
                # Keep configuration sheets visible as requested
                self._hide_config_sheets(writer)
            
            logger.info(f"Created integrated 7-sheet workbench file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to create integrated workbench file {file_path}: {str(e)}")
            return False
    
    def _create_stages_sheet(self) -> pd.DataFrame:
        """Create stages sheet structure with empty data.
        Only column headers, ready for user to populate.
        Structure matches workbench requirements - DO NOT CHANGE WITHOUT PERMISSION."""
        
        # Define column structure for stages sheet - exact order as specified by user
        stages_columns = [
            'stage_id', 
            'stage_name', 
            'platform',
            'artifact_side',
            'stage_description',
            'processing_order',
            'is_active',
            'default_artifact_type',
            'technical_fields_required',
            'partition_strategy',
            'notes'
        ]
        
        # Return empty DataFrame with just the column structure
        return pd.DataFrame(columns=stages_columns)
    
    def _create_artifacts_sheet(self) -> pd.DataFrame:
        """Create artifacts sheet structure with empty data.
        Only column headers, ready for user to populate.
        Structure matches workbench requirements - DO NOT CHANGE WITHOUT PERMISSION."""
        
        # Define column structure for artifacts sheet - exact order as specified by user
        artifacts_columns = [
            'stage_id',
            'stage_name',
            'artifact_id',
            'artifact_name',
            'artifact_type',
            'artifact_topology',
            'upstream_artifact',
            'upstream_relation',
            'relation_type',
            'artifact_relation_direction',
            'artifact_domain',
            'artifact_comment',
            'ddl_template',
            'etl_template'
        ]
        
        # Return empty DataFrame with just the column structure
        return pd.DataFrame(columns=artifacts_columns)
    
    def _create_columns_sheet(self) -> pd.DataFrame:
        """Create columns sheet structure with empty data.
        Only column headers, ready for user to populate.
        Structure matches workbench requirements - DO NOT CHANGE WITHOUT PERMISSION."""
        
        # Define column structure for columns sheet - exact order as specified
        columns_columns = [
            'stage_id',
            'stage_name',
            'artifact_id',
            'artifact_name',
            'column_id',
            'column_name',
            'data_type',
            'order',
            'column_business_name',
            'column_group',
            'column_comment'
        ]
        
        # Return empty DataFrame with just the column structure
        return pd.DataFrame(columns=columns_columns)
    
    def _apply_workbench_formatting(self, writer):
        """Apply formatting to the workbench Excel sheets.
        Includes light grey formatting for lookup columns (including headers), freeze panes, and autofilter."""
        try:
            from openpyxl.styles import PatternFill
            
            # Light grey fill for lookup columns
            light_grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Get workbooks for formatting
            workbook = writer.book
            
            # Format stages sheet
            if 'stages' in workbook.sheetnames:
                stages_sheet = workbook['stages']
                # Freeze first row
                stages_sheet.freeze_panes = 'A2'
                # Add autofilter to first row
                stages_sheet.auto_filter.ref = stages_sheet.dimensions
                # Apply light grey to stage_name column (column B) - including header
                for row in range(1, stages_sheet.max_row + 1):  # Start from row 1 to include header
                    cell = stages_sheet.cell(row=row, column=2)  # column B
                    cell.fill = light_grey_fill
            
            # Format artifacts sheet
            if 'artifacts' in workbook.sheetnames:
                artifacts_sheet = workbook['artifacts']
                # Freeze first row
                artifacts_sheet.freeze_panes = 'A2'
                # Add autofilter to first row
                artifacts_sheet.auto_filter.ref = artifacts_sheet.dimensions
                # Apply light grey to stage_name column (column B) - including header
                for row in range(1, artifacts_sheet.max_row + 1):  # Start from row 1 to include header
                    cell = artifacts_sheet.cell(row=row, column=2)  # column B
                    cell.fill = light_grey_fill
            
            # Format columns sheet
            if 'columns' in workbook.sheetnames:
                columns_sheet = workbook['columns']
                # Freeze first row
                columns_sheet.freeze_panes = 'A2'
                # Add autofilter to first row
                columns_sheet.auto_filter.ref = columns_sheet.dimensions
                # Apply light grey to lookup columns - including headers
                for row in range(1, columns_sheet.max_row + 1):  # Start from row 1 to include header
                    # stage_name column (column B)
                    cell_stage = columns_sheet.cell(row=row, column=2)  # column B
                    cell_stage.fill = light_grey_fill
                    # artifact_name column (column D)
                    cell_artifact = columns_sheet.cell(row=row, column=4)  # column D
                    cell_artifact.fill = light_grey_fill
            
            logger.info("Applied workbench formatting including freeze panes, autofilter, and light grey for lookup columns (including headers)")
            
        except ImportError:
            logger.warning("openpyxl not available for formatting")
        except Exception as e:
            logger.warning(f"Could not apply workbench formatting: {str(e)}")

    def _create_config_metadata_sheet(self, project_name: str) -> pd.DataFrame:
        """Create conf_1_stages configuration sheet based on real configuration structure."""
        stages_data = {
            'stage_id': ['s0', 's1', 's2', 's3', 's4'],
            'stage_name': ['0_drop_zone', '1_bronze', '2_silver', '3_gold', '4_mart'],
            'platform': ['databricks', 'databricks', 'databricks', 'databricks', 'databricks'],
            'artifact_side': ['source', 'source', 'source', 'source', 'consumption'],
            'description': [
                'Drop zone for raw data files',
                'Bronze layer for raw structured data',
                'Silver layer for cleaned and validated data',
                'Gold layer for business-ready data',
                'Data mart for specific business domains'
            ],
            'processing_notes': [
                'Raw data ingestion, minimal processing',
                'Raw storage with basic structure',
                'Cleaned and deduplicated source data',
                'Business logic applied, conformed dimensions',
                'Aggregated and optimized for reporting'
            ]
        }
        return pd.DataFrame(stages_data)

    def _create_config_settings_sheet(self) -> pd.DataFrame:
        """Create conf_2_technical_columns configuration sheet based on real configuration structure."""
        tech_columns_data = {
            'stage_id': ['s1', 's1', 's1', 's1', 's1', 's2', 's2', 's2', 's3', 's3', 's4'],
            'stage_name': ['1_bronze', '1_bronze', '1_bronze', '1_bronze', '1_bronze', 
                          '2_silver', '2_silver', '2_silver', '3_gold', '3_gold', '4_mart'],
            'column_name': ['__SourceSystem', '__SourceFileName', '__SourceFilePath', '__LoadDate', '__PartitionDate',
                           '__IsActive', '__ValidFrom', '__ValidTo', '__BusinessKey', '__EffectiveDate', '__LastUpdated'],
            'data_type': ['STRING', 'STRING', 'STRING', 'TIMESTAMP', 'DATE',
                         'BOOLEAN', 'TIMESTAMP', 'TIMESTAMP', 'STRING', 'TIMESTAMP', 'TIMESTAMP'],
            'include_in_tech_fields': [True, True, True, True, True, True, True, True, True, True, True],
            'take_to_next_level': [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0],
            'artifact_type_specific': ['all', 'all', 'all', 'all', 'all', 'all', 'scd2', 'scd2', 'all', 'all', 'all'],
            'description': [
                'Source system identifier', 'Original source file name', 'Source file path',
                'Load timestamp', 'Partition date for optimization', 'Record active flag',
                'Valid from timestamp', 'Valid to timestamp', 'Business key identifier',
                'Business effective date', 'Last updated timestamp'
            ]
        }
        return pd.DataFrame(tech_columns_data)

    def _create_config_validation_sheet(self) -> pd.DataFrame:
        """Create conf_3_relations configuration sheet based on real configuration structure."""
        relations_data = {
            'relation_type': ['main', 'get_key', 'lookup', 'custom'],
            'description': [
                'Full column propagation with technical fields',
                'Dimension key propagation for fact tables',
                'Limited column lookup with priority-based selection',
                'Custom relation logic defined by user'
            ],
            'processing_logic': [
                'Context-aware processing based on artifact type and stage transition',
                'Extracts surrogate keys (SKs) and business keys (BKs) only',
                'Priority order: SKs → BKs → Attributes, configurable limit',
                'User-defined transformation logic'
            ],
            'field_limit': ['No limit', 'Keys only', '3 (default)', 'User defined'],
            'use_cases': [
                'Standard column cascading between stages',
                'Foreign key relationships in fact tables',
                'Reference lookups and denormalization',
                'Special business requirements'
            ]
        }
        return pd.DataFrame(relations_data)

    def _create_config_templates_sheet(self) -> pd.DataFrame:
        """Create conf_4_data_mappings configuration sheet based on real configuration structure."""
        data_mappings_data = {
            'source': ['INT', 'BIGINT', 'SMALLINT', 'DECIMAL', 'FLOAT', 'DOUBLE', 'STRING', 'VARCHAR', 'CHAR', 'BOOLEAN', 
                      'DATE', 'TIMESTAMP', 'BINARY', 'ARRAY', 'MAP', 'STRUCT'],
            'sql_server': ['INT', 'BIGINT', 'SMALLINT', 'DECIMAL', 'FLOAT', 'FLOAT', 'NVARCHAR', 'NVARCHAR', 'NCHAR', 'BIT',
                          'DATE', 'DATETIME2', 'VARBINARY', 'NVARCHAR', 'NVARCHAR', 'NVARCHAR'],
            'databricks': ['INT', 'BIGINT', 'SMALLINT', 'DECIMAL', 'FLOAT', 'DOUBLE', 'STRING', 'STRING', 'STRING', 'BOOLEAN',
                          'DATE', 'TIMESTAMP', 'BINARY', 'ARRAY', 'MAP', 'STRUCT'],
            'power_bi': ['INT64', 'INT64', 'INT64', 'DECIMAL', 'DOUBLE', 'DOUBLE', 'TEXT', 'TEXT', 'TEXT', 'TRUE/FALSE',
                        'DATE', 'DATETIME', 'BINARY', 'TEXT', 'TEXT', 'TEXT'],
            'notes': ['Standard integer', 'Large integer', 'Small integer', 'Decimal number', 'Single precision float', 
                     'Double precision float', 'Text string', 'Variable length string', 'Fixed length string', 'Boolean flag',
                     'Date only', 'Date and time', 'Binary data', 'Array collection', 'Key-value pairs', 'Structured data']
        }
        return pd.DataFrame(data_mappings_data)

    def _hide_config_sheets(self, writer):
        """Do not hide the configuration sheets - keep them visible as requested."""
        # Configuration sheets should remain visible as per user request
        logger.info("Configuration sheets kept visible as requested")


def create_default_workbench(output_path: str = "workbench_default.xlsx") -> bool:
    """
    Convenience function to create a default workbench file.
    
    Args:
        output_path (str): Path for the output file
        
    Returns:
        bool: True if successful, False otherwise
    """
    manager = WorkbenchSetupManager()
    return manager.create_default_workbench_file(output_path)


if __name__ == "__main__":
    # Example usage
    print("Creating default workbench file...")
    success = create_default_workbench("workbench_example.xlsx")
    if success:
        print("✅ Default workbench file created successfully!")
    else:
        print("❌ Failed to create default workbench file")