# Test Configuration for DWH Creator
"""
pytest configuration and fixtures for DWH Creator testing.
"""

import pytest
import os
import tempfile
import shutil
from pathlib import Path
import pandas as pd
import openpyxl
from unittest.mock import Mock, patch

# Add src to Python path for imports
import sys
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

@pytest.fixture(scope="session")
def project_root_path():
    """Return the project root directory path."""
    return Path(__file__).parent.parent

@pytest.fixture(scope="function")
def temp_dir():
    """Create a temporary directory for tests."""
    temp_path = tempfile.mkdtemp()
    yield Path(temp_path)
    shutil.rmtree(temp_path, ignore_errors=True)

@pytest.fixture(scope="function")
def sample_project_dir(temp_dir):
    """Create a sample project directory structure."""
    project_path = temp_dir / "test_project"
    project_path.mkdir()
    
    # Create subdirectories
    (project_path / "data").mkdir()
    (project_path / "logs").mkdir()
    (project_path / "outputs").mkdir()
    
    # Create sample files
    (project_path / "README.md").write_text("# Test Project")
    
    return project_path

@pytest.fixture(scope="function")
def sample_workbook_path(sample_project_dir):
    """Create a sample Excel workbook with test data."""
    workbook_path = sample_project_dir / "data" / "test_workbook.xlsx"
    
    # Create workbook with sample data
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Stages sheet
    stages_ws = wb.create_sheet("Stages")
    stages_data = [
        ["Stage ID", "Stage Name", "Stage Description", "Platform", "Source/Business Side"],
        [1, "bronze", "Raw data ingestion", "Databricks", "Source"],
        [2, "silver", "Cleaned data", "Databricks", "Source"],
        [3, "gold", "Business-ready data", "Databricks", "Business"]
    ]
    for row in stages_data:
        stages_ws.append(row)
    
    # Create Artifacts sheet
    artifacts_ws = wb.create_sheet("Artifacts")
    artifacts_data = [
        ["Artifact ID", "Artifact Name", "Stage Name", "Artifact Comment"],
        [1, "customers_bronze", "bronze", "Raw customer data"],
        [2, "orders_bronze", "bronze", "Raw order data"],
        [3, "customers_silver", "silver", "Cleaned customer data"]
    ]
    for row in artifacts_data:
        artifacts_ws.append(row)
    
    # Create Columns sheet
    columns_ws = wb.create_sheet("Columns")
    columns_data = [
        ["Column ID", "Column Name", "Artifact ID", "Data Type", "Column Business Name", "Column Comment"],
        [1, "customer_id", 1, "INT", "customer_id", "Customer identifier"],
        [2, "first_name", 1, "VARCHAR(50)", "customer_first_name", "Customer first name"],
        [3, "order_id", 2, "INT", "order_id", "Order identifier"],
        [4, "order_date", 2, "DATE", "order_date", "Order placement date"]
    ]
    for row in columns_data:
        columns_ws.append(row)
    
    wb.save(workbook_path)
    return workbook_path

@pytest.fixture(scope="function")
def mock_openai_client():
    """Create a mock OpenAI client for testing AI features."""
    with patch('utils.ai_comment_generator.OpenAI') as mock_openai_class:
        # Mock successful response
        mock_response = Mock()
        mock_response.choices = [Mock()]
        mock_response.choices[0].message.content = "Test AI response"
        
        # Mock the client instance
        mock_client_instance = Mock()
        mock_client_instance.chat.completions.create.return_value = mock_response
        mock_openai_class.return_value = mock_client_instance
        
        yield mock_openai_class

@pytest.fixture(scope="function")
def mock_config():
    """Create mock configuration for testing."""
    # Don't use real environment variables in tests
    mock_env = {
        'OPENAI_API_KEY': '',  # Empty to simulate no API key
        'OPENAI_MODEL': 'gpt-3.5-turbo',
        'LOG_LEVEL': 'INFO'
    }
    with patch.dict(os.environ, mock_env, clear=False):
        with patch('utils.app_config.AppConfig.get_openai_api_key', return_value=None):
            yield

@pytest.fixture(scope="function")
def sample_dataframes():
    """Create sample pandas DataFrames for testing."""
    stages_df = pd.DataFrame({
        'Stage ID': [1, 2, 3],
        'Stage Name': ['bronze', 'silver', 'gold'],
        'Stage Description': ['Raw data', 'Cleaned data', 'Business data'],
        'Platform': ['Databricks', 'Databricks', 'Databricks'],
        'Source/Business Side': ['Source', 'Source', 'Business']
    })
    
    artifacts_df = pd.DataFrame({
        'Artifact ID': [1, 2, 3],
        'Artifact Name': ['customers_bronze', 'orders_bronze', 'customers_silver'],
        'Stage Name': ['bronze', 'bronze', 'silver'],
        'Artifact Comment': ['Raw customer data', 'Raw order data', 'Cleaned customer data']
    })
    
    columns_df = pd.DataFrame({
        'Column ID': [1, 2, 3, 4],
        'Column Name': ['customer_id', 'first_name', 'order_id', 'order_date'],
        'Artifact ID': [1, 1, 2, 2],
        'Data Type': ['INT', 'VARCHAR(50)', 'INT', 'DATE'],
        'Column Business Name': ['customer_id', 'customer_first_name', 'order_id', 'order_date'],
        'Column Comment': ['Customer identifier', 'Customer first name', 'Order identifier', 'Order placement date']
    })
    
    return {
        'stages': stages_df,
        'artifacts': artifacts_df,
        'columns': columns_df
    }

# Test markers
pytest_markers = [
    "unit: Unit tests for individual functions/classes",
    "integration: Integration tests for component interactions", 
    "ai: Tests requiring AI/OpenAI functionality",
    "excel: Tests requiring Excel file operations",
    "slow: Tests that take longer to run",
    "network: Tests that require network connectivity"
]

def pytest_configure(config):
    """Configure pytest with custom markers."""
    for marker in pytest_markers:
        config.addinivalue_line("markers", marker)
